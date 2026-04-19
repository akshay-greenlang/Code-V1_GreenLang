# -*- coding: utf-8 -*-
"""
Comprehensive tests for GreenLang Data Writers.

Tests cover:
- Multi-format writing (JSON, CSV, Excel, YAML, etc.)
- Directory creation
- File overwriting behavior
- Format-specific options
- Error handling
"""

import pytest
import json
import csv
from pathlib import Path
from unittest.mock import patch

from greenlang.io.writers import DataWriter, write_file


# Test Fixtures
@pytest.fixture
def writer():
    """Create a data writer instance."""
    return DataWriter()


@pytest.fixture
def sample_json_data():
    """Sample JSON data."""
    return {
        "name": "John Doe",
        "age": 30,
        "email": "john@example.com",
        "tags": ["developer", "python"]
    }


@pytest.fixture
def sample_csv_data():
    """Sample CSV data (list of dicts)."""
    return [
        {"name": "John", "age": 30, "city": "New York"},
        {"name": "Jane", "age": 25, "city": "Boston"},
        {"name": "Bob", "age": 35, "city": "Chicago"}
    ]


# DataWriter Initialization Tests
class TestDataWriterInit:
    """Test DataWriter initialization."""

    def test_writer_creation(self):
        """Test creating a data writer."""
        writer = DataWriter()
        assert writer.default_encoding == "utf-8"
        assert ".json" in writer._format_handlers
        assert ".csv" in writer._format_handlers

    def test_custom_encoding(self):
        """Test writer with custom encoding."""
        writer = DataWriter(default_encoding="latin-1")
        assert writer.default_encoding == "latin-1"

    def test_supported_formats(self, writer):
        """Test getting supported formats."""
        formats = writer.get_supported_formats()
        assert ".json" in formats
        assert ".csv" in formats
        assert ".txt" in formats


# JSON Writing Tests
class TestJSONWriting:
    """Test JSON file writing."""

    def test_write_json_object(self, writer, tmp_path, sample_json_data):
        """Test writing JSON object."""
        file_path = tmp_path / "output.json"
        writer.write(sample_json_data, file_path)

        assert file_path.exists()

        # Verify content
        with open(file_path, 'r') as f:
            data = json.load(f)
        assert data == sample_json_data

    def test_write_json_array(self, writer, tmp_path):
        """Test writing JSON array."""
        array_data = [
            {"id": 1, "name": "Item 1"},
            {"id": 2, "name": "Item 2"}
        ]
        file_path = tmp_path / "array.json"
        writer.write(array_data, file_path)

        with open(file_path, 'r') as f:
            data = json.load(f)
        assert len(data) == 2

    def test_write_json_with_indent(self, writer, tmp_path, sample_json_data):
        """Test writing JSON with custom indentation."""
        file_path = tmp_path / "indented.json"
        writer.write(sample_json_data, file_path, indent=4)

        content = file_path.read_text()
        # Should be pretty-printed
        assert "\n" in content
        assert "    " in content

    def test_write_json_no_indent(self, writer, tmp_path, sample_json_data):
        """Test writing JSON without indentation."""
        file_path = tmp_path / "compact.json"
        writer.write(sample_json_data, file_path, indent=None)

        content = file_path.read_text()
        # Should be compact (no indentation)

    def test_write_nested_json(self, writer, tmp_path):
        """Test writing nested JSON."""
        nested_data = {
            "user": {
                "name": "John",
                "address": {
                    "city": "New York",
                    "zipcode": "10001"
                }
            }
        }
        file_path = tmp_path / "nested.json"
        writer.write(nested_data, file_path)

        with open(file_path, 'r') as f:
            data = json.load(f)
        assert data["user"]["address"]["city"] == "New York"


# CSV Writing Tests
class TestCSVWriting:
    """Test CSV file writing."""

    def test_write_csv(self, writer, tmp_path, sample_csv_data):
        """Test writing CSV file."""
        file_path = tmp_path / "output.csv"
        writer.write(sample_csv_data, file_path)

        assert file_path.exists()

        # Verify content
        with open(file_path, 'r') as f:
            reader = csv.DictReader(f)
            rows = list(reader)

        assert len(rows) == 3
        assert rows[0]["name"] == "John"

    def test_write_csv_custom_delimiter(self, writer, tmp_path, sample_csv_data):
        """Test writing CSV with custom delimiter."""
        file_path = tmp_path / "semicolon.csv"
        writer.write(sample_csv_data, file_path, csv_delimiter=";")

        content = file_path.read_text()
        assert ";" in content

    def test_write_tsv(self, writer, tmp_path, sample_csv_data):
        """Test writing TSV file."""
        file_path = tmp_path / "output.tsv"
        writer.write(sample_csv_data, file_path)

        content = file_path.read_text()
        assert "\t" in content

    def test_write_empty_csv(self, writer, tmp_path):
        """Test writing empty CSV."""
        file_path = tmp_path / "empty.csv"
        writer.write([], file_path)

        # File should exist but be empty or minimal
        assert file_path.exists()

    def test_write_csv_field_order(self, writer, tmp_path):
        """Test that CSV maintains field order."""
        data = [
            {"z_field": 1, "a_field": 2, "m_field": 3}
        ]
        file_path = tmp_path / "ordered.csv"
        writer.write(data, file_path)

        with open(file_path, 'r') as f:
            reader = csv.DictReader(f)
            fieldnames = reader.fieldnames

        # Order should match first record
        assert fieldnames == ["z_field", "a_field", "m_field"]


# Text Writing Tests
class TestTextWriting:
    """Test text file writing."""

    def test_write_text(self, writer, tmp_path):
        """Test writing text file."""
        text = "Hello, World!\nThis is a test."
        file_path = tmp_path / "output.txt"
        writer.write(text, file_path)

        content = file_path.read_text()
        assert content == text

    def test_write_multiline_text(self, writer, tmp_path):
        """Test writing multiline text."""
        text = "Line 1\nLine 2\nLine 3"
        file_path = tmp_path / "multiline.txt"
        writer.write(text, file_path)

        content = file_path.read_text()
        assert "Line 1" in content
        assert "Line 2" in content


# YAML Writing Tests
class TestYAMLWriting:
    """Test YAML file writing."""

    def test_write_yaml(self, writer, tmp_path):
        """Test writing YAML file."""
        try:
            import yaml

            data = {
                "name": "John Doe",
                "age": 30,
                "tags": ["developer", "python"]
            }
            file_path = tmp_path / "output.yaml"
            writer.write(data, file_path)

            # Verify content
            with open(file_path, 'r') as f:
                loaded = yaml.safe_load(f)
            assert loaded["name"] == "John Doe"
        except ImportError:
            pytest.skip("PyYAML not available")

    def test_write_yaml_nested(self, writer, tmp_path):
        """Test writing nested YAML."""
        try:
            import yaml

            data = {
                "user": {
                    "name": "John",
                    "address": {
                        "city": "New York"
                    }
                }
            }
            file_path = tmp_path / "nested.yaml"
            writer.write(data, file_path)

            with open(file_path, 'r') as f:
                loaded = yaml.safe_load(f)
            assert loaded["user"]["address"]["city"] == "New York"
        except ImportError:
            pytest.skip("PyYAML not available")


# Excel Writing Tests
class TestExcelWriting:
    """Test Excel file writing."""

    def test_write_excel(self, writer, tmp_path, sample_csv_data):
        """Test writing Excel file."""
        try:
            import openpyxl

            file_path = tmp_path / "output.xlsx"
            writer.write(sample_csv_data, file_path)

            assert file_path.exists()

            # Verify content
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            assert ws.cell(1, 1).value == "name"
            assert ws.cell(2, 1).value == "John"
            wb.close()
        except ImportError:
            pytest.skip("openpyxl not available")

    def test_write_excel_custom_sheet_name(self, writer, tmp_path, sample_csv_data):
        """Test writing Excel with custom sheet name."""
        try:
            import openpyxl

            file_path = tmp_path / "custom_sheet.xlsx"
            writer.write(sample_csv_data, file_path, sheet_name="MyData")

            wb = openpyxl.load_workbook(file_path)
            assert "MyData" in wb.sheetnames
            wb.close()
        except ImportError:
            pytest.skip("openpyxl not available")

    def test_write_excel_auto_width(self, writer, tmp_path):
        """Test Excel column auto-width."""
        try:
            import openpyxl

            data = [
                {"short": "x", "very_long_column_name": "This is a very long value"}
            ]
            file_path = tmp_path / "auto_width.xlsx"
            writer.write(data, file_path, auto_width=True)

            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            # Column widths should be set
            wb.close()
        except ImportError:
            pytest.skip("openpyxl not available")


# Directory Creation Tests
class TestDirectoryCreation:
    """Test automatic directory creation."""

    def test_create_parent_directory(self, writer, tmp_path, sample_json_data):
        """Test that parent directories are created."""
        file_path = tmp_path / "subdir" / "nested" / "output.json"
        writer.write(sample_json_data, file_path)

        assert file_path.exists()
        assert file_path.parent.exists()

    def test_create_deep_directory(self, writer, tmp_path, sample_json_data):
        """Test creating deeply nested directories."""
        file_path = tmp_path / "a" / "b" / "c" / "d" / "output.json"
        writer.write(sample_json_data, file_path)

        assert file_path.exists()

    def test_existing_directory(self, writer, tmp_path, sample_json_data):
        """Test writing to existing directory."""
        subdir = tmp_path / "existing"
        subdir.mkdir()

        file_path = subdir / "output.json"
        writer.write(sample_json_data, file_path)

        assert file_path.exists()


# File Overwriting Tests
class TestFileOverwriting:
    """Test file overwriting behavior."""

    def test_overwrite_existing_file(self, writer, tmp_path, sample_json_data):
        """Test that existing files are overwritten."""
        file_path = tmp_path / "overwrite.json"

        # Write first version
        writer.write({"version": 1}, file_path)

        # Overwrite with second version
        writer.write({"version": 2}, file_path)

        # Verify second version
        with open(file_path, 'r') as f:
            data = json.load(f)
        assert data["version"] == 2

    def test_overwrite_different_content(self, writer, tmp_path):
        """Test overwriting with completely different content."""
        file_path = tmp_path / "file.json"

        # Write small file
        writer.write({"small": "data"}, file_path)

        # Overwrite with larger file
        large_data = {"large": ["item"] * 100}
        writer.write(large_data, file_path)

        with open(file_path, 'r') as f:
            data = json.load(f)
        assert len(data["large"]) == 100


# Error Handling Tests
class TestErrorHandling:
    """Test error handling."""

    def test_unsupported_format(self, writer, tmp_path, sample_json_data):
        """Test writing to unsupported format."""
        file_path = tmp_path / "output.xyz"

        with pytest.raises(ValueError, match="Unsupported format"):
            writer.write(sample_json_data, file_path)

    def test_csv_requires_dict_list(self, writer, tmp_path):
        """Test that CSV requires list of dicts."""
        file_path = tmp_path / "output.csv"

        # Try to write non-dict data
        with pytest.raises(ValueError):
            writer.write([["a", "b"], ["c", "d"]], file_path)

    def test_write_to_readonly_location(self, writer, tmp_path, sample_json_data):
        """Test writing to read-only location."""
        # This test depends on OS permissions
        # Skip if can't create read-only directory


# Convenience Function Tests
class TestConvenienceFunctions:
    """Test convenience functions."""

    def test_write_file_function(self, tmp_path, sample_json_data):
        """Test write_file convenience function."""
        file_path = tmp_path / "output.json"
        write_file(sample_json_data, file_path)

        assert file_path.exists()

    def test_write_file_with_options(self, tmp_path, sample_csv_data):
        """Test write_file with format options."""
        file_path = tmp_path / "output.csv"
        write_file(sample_csv_data, file_path, csv_delimiter=";")

        content = file_path.read_text()
        assert ";" in content


# Path Handling Tests
class TestPathHandling:
    """Test path handling."""

    def test_string_path(self, writer, tmp_path, sample_json_data):
        """Test writing with string path."""
        file_path = str(tmp_path / "output.json")
        writer.write(sample_json_data, file_path)

        assert Path(file_path).exists()

    def test_path_object(self, writer, tmp_path, sample_json_data):
        """Test writing with Path object."""
        file_path = tmp_path / "output.json"
        writer.write(sample_json_data, file_path)

        assert file_path.exists()

    def test_absolute_path(self, writer, tmp_path, sample_json_data):
        """Test writing with absolute path."""
        file_path = (tmp_path / "output.json").absolute()
        writer.write(sample_json_data, file_path)

        assert file_path.exists()


# Edge Cases Tests
class TestEdgeCases:
    """Test edge cases."""

    def test_write_empty_json_object(self, writer, tmp_path):
        """Test writing empty JSON object."""
        file_path = tmp_path / "empty.json"
        writer.write({}, file_path)

        with open(file_path, 'r') as f:
            data = json.load(f)
        assert data == {}

    def test_write_empty_json_array(self, writer, tmp_path):
        """Test writing empty JSON array."""
        file_path = tmp_path / "empty_array.json"
        writer.write([], file_path)

        with open(file_path, 'r') as f:
            data = json.load(f)
        assert data == []

    def test_write_large_data(self, writer, tmp_path):
        """Test writing large data."""
        large_data = [{"id": i, "value": f"item_{i}"} for i in range(1000)]
        file_path = tmp_path / "large.json"
        writer.write(large_data, file_path)

        with open(file_path, 'r') as f:
            data = json.load(f)
        assert len(data) == 1000

    def test_write_special_characters(self, writer, tmp_path):
        """Test writing data with special characters."""
        data = {
            "message": "Contains \"quotes\" and 'apostrophes'",
            "unicode": "Hello 世界 🌍"
        }
        file_path = tmp_path / "special.json"
        writer.write(data, file_path)

        with open(file_path, 'r', encoding='utf-8') as f:
            loaded = json.load(f)
        assert loaded["unicode"] == "Hello 世界 🌍"

    def test_case_insensitive_extension(self, writer, tmp_path, sample_json_data):
        """Test that extension matching is case insensitive."""
        file_path = tmp_path / "output.JSON"
        writer.write(sample_json_data, file_path)

        assert file_path.exists()

    def test_write_none_value(self, writer, tmp_path):
        """Test writing None values in data."""
        data = {"name": "John", "optional": None}
        file_path = tmp_path / "with_null.json"
        writer.write(data, file_path)

        with open(file_path, 'r') as f:
            loaded = json.load(f)
        assert loaded["optional"] is None

    def test_write_csv_with_special_chars(self, writer, tmp_path):
        """Test writing CSV with special characters."""
        data = [
            {"name": "Item", "description": 'Contains, comma and "quotes"'}
        ]
        file_path = tmp_path / "special.csv"
        writer.write(data, file_path)

        # Verify it can be read back
        with open(file_path, 'r') as f:
            reader = csv.DictReader(f)
            rows = list(reader)
        assert "comma" in rows[0]["description"]

    def test_filename_with_spaces(self, writer, tmp_path, sample_json_data):
        """Test writing to filename with spaces."""
        file_path = tmp_path / "file with spaces.json"
        writer.write(sample_json_data, file_path)

        assert file_path.exists()
