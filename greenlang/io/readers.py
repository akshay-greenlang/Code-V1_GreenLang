"""
GreenLang Data Readers
Multi-format data reading with automatic format detection.
"""

from typing import Any, Dict, List, Optional, Union
from pathlib import Path
import logging
import json
import csv

logger = logging.getLogger(__name__)


class DataReader:
    """
    Multi-format data reader with automatic format detection.

    Supported formats:
    - JSON (.json)
    - CSV (.csv)
    - TSV (.tsv)
    - TXT (.txt)
    - YAML (.yaml, .yml) - if PyYAML available
    - Excel (.xlsx, .xls) - if openpyxl/xlrd available
    - Parquet (.parquet) - if pyarrow available
    - XML (.xml) - if lxml available

    Example:
        reader = DataReader()
        data = reader.read("data.json")

        # With options
        data = reader.read("data.csv", csv_delimiter=";", csv_has_header=True)
    """

    def __init__(self, default_encoding: str = "utf-8"):
        """
        Initialize data reader.

        Args:
            default_encoding: Default text encoding
        """
        self.default_encoding = default_encoding
        self._format_handlers = {
            ".json": self._read_json,
            ".csv": self._read_csv,
            ".tsv": self._read_tsv,
            ".txt": self._read_txt,
        }

        # Check for optional dependencies and register handlers
        try:
            import yaml
            self._format_handlers[".yaml"] = self._read_yaml
            self._format_handlers[".yml"] = self._read_yaml
        except ImportError:
            logger.debug("PyYAML not available, YAML support disabled")

        try:
            import openpyxl
            self._format_handlers[".xlsx"] = self._read_excel
        except ImportError:
            logger.debug("openpyxl not available, Excel (.xlsx) support disabled")

        try:
            import xlrd
            self._format_handlers[".xls"] = self._read_excel_legacy
        except ImportError:
            logger.debug("xlrd not available, Excel (.xls) support disabled")

        try:
            import pyarrow.parquet as pq
            self._format_handlers[".parquet"] = self._read_parquet
        except ImportError:
            logger.debug("pyarrow not available, Parquet support disabled")

        try:
            from lxml import etree
            self._format_handlers[".xml"] = self._read_xml
        except ImportError:
            logger.debug("lxml not available, XML support disabled")

    def read(self, file_path: Union[str, Path], **kwargs) -> Any:
        """
        Read data from file with automatic format detection.

        Args:
            file_path: Path to file
            **kwargs: Format-specific options

        Returns:
            Loaded data (format depends on file type)

        Raises:
            FileNotFoundError: If file doesn't exist
            ValueError: If format not supported
        """
        path = Path(file_path)

        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        # Determine format from extension
        ext = path.suffix.lower()

        if ext not in self._format_handlers:
            raise ValueError(f"Unsupported format: {ext}")

        # Read using appropriate handler
        logger.debug(f"Reading {path} as {ext} format")
        handler = self._format_handlers[ext]
        return handler(path, **kwargs)

    def _read_json(self, path: Path, **kwargs) -> Union[Dict, List]:
        """Read JSON file."""
        with open(path, 'r', encoding=self.default_encoding) as f:
            return json.load(f)

    def _read_csv(self, path: Path, **kwargs) -> List[Dict[str, Any]]:
        """Read CSV file."""
        delimiter = kwargs.get("csv_delimiter", ",")
        has_header = kwargs.get("csv_has_header", True)

        records = []

        with open(path, 'r', encoding=self.default_encoding, newline='') as f:
            if has_header:
                reader = csv.DictReader(f, delimiter=delimiter)
            else:
                reader = csv.reader(f, delimiter=delimiter)

            for row in reader:
                if has_header:
                    records.append(dict(row))
                else:
                    records.append(list(row))

        return records

    def _read_tsv(self, path: Path, **kwargs) -> List[Dict[str, Any]]:
        """Read TSV file (tab-separated)."""
        kwargs["csv_delimiter"] = "\t"
        return self._read_csv(path, **kwargs)

    def _read_txt(self, path: Path, **kwargs) -> str:
        """Read text file."""
        with open(path, 'r', encoding=self.default_encoding) as f:
            return f.read()

    def _read_yaml(self, path: Path, **kwargs) -> Any:
        """Read YAML file."""
        import yaml

        with open(path, 'r', encoding=self.default_encoding) as f:
            return yaml.safe_load(f)

    def _read_excel(self, path: Path, **kwargs) -> List[Dict[str, Any]]:
        """Read Excel file (.xlsx) using openpyxl."""
        import openpyxl

        sheet_name = kwargs.get("sheet_name", 0)

        workbook = openpyxl.load_workbook(path, read_only=True)

        if isinstance(sheet_name, int):
            sheet = workbook.worksheets[sheet_name]
        else:
            sheet = workbook[sheet_name]

        # Get headers from first row
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

        # Read data rows
        records = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            record = {headers[i]: row[i] for i in range(len(headers))}
            records.append(record)

        workbook.close()
        return records

    def _read_excel_legacy(self, path: Path, **kwargs) -> List[Dict[str, Any]]:
        """Read Excel file (.xls) using xlrd."""
        import xlrd

        sheet_index = kwargs.get("sheet_name", 0)

        workbook = xlrd.open_workbook(path)
        sheet = workbook.sheet_by_index(sheet_index) if isinstance(sheet_index, int) else workbook.sheet_by_name(sheet_index)

        # Get headers
        headers = [sheet.cell_value(0, col) for col in range(sheet.ncols)]

        # Read data
        records = []
        for row_idx in range(1, sheet.nrows):
            record = {headers[col]: sheet.cell_value(row_idx, col) for col in range(sheet.ncols)}
            records.append(record)

        return records

    def _read_parquet(self, path: Path, **kwargs) -> List[Dict[str, Any]]:
        """Read Parquet file."""
        import pyarrow.parquet as pq

        table = pq.read_table(path)
        df = table.to_pandas()
        return df.to_dict('records')

    def _read_xml(self, path: Path, **kwargs) -> Any:
        """Read XML file."""
        from lxml import etree

        tree = etree.parse(str(path))
        root = tree.getroot()

        # Simple conversion to dict
        def element_to_dict(element):
            result = {element.tag: {}}

            # Add attributes
            if element.attrib:
                result[element.tag]["@attributes"] = element.attrib

            # Add text
            if element.text and element.text.strip():
                result[element.tag]["text"] = element.text.strip()

            # Add children
            for child in element:
                child_dict = element_to_dict(child)
                result[element.tag].update(child_dict)

            return result

        return element_to_dict(root)

    def get_supported_formats(self) -> List[str]:
        """Get list of supported file formats."""
        return list(self._format_handlers.keys())


def read_file(file_path: Union[str, Path], **kwargs) -> Any:
    """
    Convenience function to read a file.

    Args:
        file_path: Path to file
        **kwargs: Format-specific options

    Returns:
        Loaded data
    """
    reader = DataReader()
    return reader.read(file_path, **kwargs)
