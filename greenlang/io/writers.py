"""
GreenLang Data Writers
Multi-format data writing with format conversion.
"""

from typing import Any, Dict, List, Optional, Union
from pathlib import Path
import logging
import json
import csv

logger = logging.getLogger(__name__)


class DataWriter:
    """
    Multi-format data writer.

    Supported formats:
    - JSON (.json)
    - CSV (.csv)
    - TSV (.tsv)
    - TXT (.txt)
    - YAML (.yaml, .yml) - if PyYAML available
    - Excel (.xlsx) - if openpyxl available
    - Parquet (.parquet) - if pyarrow available

    Example:
        writer = DataWriter()
        writer.write(data, "output.json", indent=2)

        # CSV with custom delimiter
        writer.write(records, "output.csv", csv_delimiter=";")
    """

    def __init__(self, default_encoding: str = "utf-8"):
        """
        Initialize data writer.

        Args:
            default_encoding: Default text encoding
        """
        self.default_encoding = default_encoding
        self._format_handlers = {
            ".json": self._write_json,
            ".csv": self._write_csv,
            ".tsv": self._write_tsv,
            ".txt": self._write_txt,
        }

        # Check for optional dependencies
        try:
            import yaml
            self._format_handlers[".yaml"] = self._write_yaml
            self._format_handlers[".yml"] = self._write_yaml
        except ImportError:
            logger.debug("PyYAML not available, YAML writing disabled")

        try:
            import openpyxl
            self._format_handlers[".xlsx"] = self._write_excel
        except ImportError:
            logger.debug("openpyxl not available, Excel writing disabled")

        try:
            import pyarrow.parquet as pq
            self._format_handlers[".parquet"] = self._write_parquet
        except ImportError:
            logger.debug("pyarrow not available, Parquet writing disabled")

    def write(self, data: Any, file_path: Union[str, Path], **kwargs):
        """
        Write data to file with automatic format detection.

        Args:
            data: Data to write
            file_path: Path to output file
            **kwargs: Format-specific options

        Raises:
            ValueError: If format not supported
        """
        path = Path(file_path)

        # Create parent directory if needed
        path.parent.mkdir(parents=True, exist_ok=True)

        # Determine format from extension
        ext = path.suffix.lower()

        if ext not in self._format_handlers:
            raise ValueError(f"Unsupported format: {ext}")

        # Write using appropriate handler
        logger.debug(f"Writing to {path} as {ext} format")
        handler = self._format_handlers[ext]
        handler(data, path, **kwargs)

    def _write_json(self, data: Any, path: Path, **kwargs):
        """Write JSON file."""
        indent = kwargs.get("indent", 2)

        with open(path, 'w', encoding=self.default_encoding) as f:
            json.dump(data, f, indent=indent)

    def _write_csv(self, data: List[Dict[str, Any]], path: Path, **kwargs):
        """Write CSV file."""
        if not data:
            logger.warning("No data to write to CSV")
            return

        delimiter = kwargs.get("csv_delimiter", ",")

        # Get fieldnames from first record
        if isinstance(data[0], dict):
            fieldnames = list(data[0].keys())
        else:
            raise ValueError("CSV writing requires list of dictionaries")

        with open(path, 'w', encoding=self.default_encoding, newline='') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter=delimiter)
            writer.writeheader()
            writer.writerows(data)

    def _write_tsv(self, data: List[Dict[str, Any]], path: Path, **kwargs):
        """Write TSV file (tab-separated)."""
        kwargs["csv_delimiter"] = "\t"
        self._write_csv(data, path, **kwargs)

    def _write_txt(self, data: str, path: Path, **kwargs):
        """Write text file."""
        with open(path, 'w', encoding=self.default_encoding) as f:
            f.write(str(data))

    def _write_yaml(self, data: Any, path: Path, **kwargs):
        """Write YAML file."""
        import yaml

        with open(path, 'w', encoding=self.default_encoding) as f:
            yaml.dump(data, f, default_flow_style=False)

    def _write_excel(self, data: List[Dict[str, Any]], path: Path, **kwargs):
        """Write Excel file (.xlsx) using openpyxl."""
        import openpyxl
        from openpyxl.utils import get_column_letter

        if not data:
            logger.warning("No data to write to Excel")
            return

        sheet_name = kwargs.get("sheet_name", "Sheet1")

        # Create workbook
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheet_name

        # Get fieldnames
        fieldnames = list(data[0].keys())

        # Write headers
        for col_idx, fieldname in enumerate(fieldnames, start=1):
            sheet.cell(row=1, column=col_idx, value=fieldname)

        # Write data rows
        for row_idx, record in enumerate(data, start=2):
            for col_idx, fieldname in enumerate(fieldnames, start=1):
                value = record.get(fieldname)
                sheet.cell(row=row_idx, column=col_idx, value=value)

        # Auto-size columns (optional)
        if kwargs.get("auto_width", True):
            for col_idx, fieldname in enumerate(fieldnames, start=1):
                max_length = len(str(fieldname))
                for record in data:
                    value_length = len(str(record.get(fieldname, "")))
                    max_length = max(max_length, value_length)

                sheet.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)

        workbook.save(path)
        workbook.close()

    def _write_parquet(self, data: List[Dict[str, Any]], path: Path, **kwargs):
        """Write Parquet file."""
        import pyarrow as pa
        import pyarrow.parquet as pq

        # Convert to PyArrow table
        table = pa.Table.from_pylist(data)

        # Write to file
        pq.write_table(table, path)

    def get_supported_formats(self) -> List[str]:
        """Get list of supported file formats."""
        return list(self._format_handlers.keys())


def write_file(data: Any, file_path: Union[str, Path], **kwargs):
    """
    Convenience function to write data to file.

    Args:
        data: Data to write
        file_path: Path to output file
        **kwargs: Format-specific options
    """
    writer = DataWriter()
    writer.write(data, file_path, **kwargs)
