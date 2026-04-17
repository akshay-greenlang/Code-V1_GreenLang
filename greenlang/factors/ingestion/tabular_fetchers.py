# -*- coding: utf-8 -*-
"""Tabular upstream helpers (D2): CSV and XLSX bytes → rows (no network)."""

from __future__ import annotations

import csv
import io
from typing import Any, Dict, List, Optional


def parse_csv_bytes(data: bytes, encoding: str = "utf-8") -> List[Dict[str, Any]]:
    """Parse CSV to list of dict rows (header row required)."""
    text = data.decode(encoding, errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader]


def parse_xlsx_bytes(data: bytes, sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
    """
    Parse first sheet (or named sheet) to dict rows using openpyxl when installed.
    """
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("openpyxl is required for XLSX parsing") from exc

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c or "").strip() for c in rows[0]]
    out: List[Dict[str, Any]] = []
    for r in rows[1:]:
        if r is None or all(v is None for v in r):
            continue
        out.append({header[i]: r[i] for i in range(min(len(header), len(r)))})
    return out
