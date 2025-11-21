# -*- coding: utf-8 -*-
"""
GreenLang Base Reporter Agent
Specialized base class for reporting and data aggregation.
"""

from typing import Any, Dict, List, Optional, Union
from pydantic import BaseModel, Field
from abc import abstractmethod
import logging
from datetime import datetime
from collections import defaultdict

from .base import BaseAgent, AgentConfig, AgentResult
from greenlang.determinism import DeterministicClock

logger = logging.getLogger(__name__)


class ReporterConfig(AgentConfig):
    """Configuration for reporter agents."""
    output_format: str = Field(default="markdown", description="Output format (markdown, html, json, excel)")
    include_summary: bool = Field(default=True, description="Include summary section")
    include_details: bool = Field(default=True, description="Include detailed sections")
    include_charts: bool = Field(default=False, description="Include charts/visualizations")
    template_path: Optional[str] = Field(default=None, description="Optional custom template path")


class ReportSection(BaseModel):
    """A section in the report."""
    title: str = Field(..., description="Section title")
    content: Union[str, List[Dict[str, Any]], Dict[str, Any]] = Field(..., description="Section content")
    level: int = Field(default=2, description="Heading level (1-6)")
    section_type: str = Field(default="text", description="Section type (text, table, chart, list)")


class BaseReporter(BaseAgent):
    """
    Base class for reporting agents.

    Provides:
    - Multi-format output (Markdown, HTML, JSON, Excel)
    - Data aggregation utilities
    - Template-based reporting
    - Summary generation
    - Section management

    Example:
        class SalesReporter(BaseReporter):
            def aggregate_data(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
                # Aggregate sales data
                return {
                    "total_sales": sum(input_data['sales']),
                    "average": mean(input_data['sales'])
                }

            def build_sections(self, aggregated_data: Dict[str, Any]) -> List[ReportSection]:
                return [
                    ReportSection(title="Summary", content=f"Total: {aggregated_data['total_sales']}"),
                    ReportSection(title="Details", content=aggregated_data, section_type="table")
                ]
    """

    def __init__(self, config: Optional[ReporterConfig] = None):
        """Initialize reporter with configuration."""
        if config is None:
            config = ReporterConfig(
                name=self.__class__.__name__,
                description=self.__class__.__doc__ or "Reporter agent"
            )
        super().__init__(config)
        self.config: ReporterConfig = config

        # Report state
        self.sections: List[ReportSection] = []
        self.metadata: Dict[str, Any] = {}

    @abstractmethod
    def aggregate_data(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Aggregate input data for reporting.
        Must be implemented by subclasses.

        Args:
            input_data: Raw input data

        Returns:
            Aggregated data dictionary
        """
        pass

    @abstractmethod
    def build_sections(self, aggregated_data: Dict[str, Any]) -> List[ReportSection]:
        """
        Build report sections from aggregated data.
        Must be implemented by subclasses.

        Args:
            aggregated_data: Aggregated data

        Returns:
            List of ReportSection objects
        """
        pass

    def add_section(
        self,
        title: str,
        content: Any,
        level: int = 2,
        section_type: str = "text"
    ):
        """
        Add a section to the report.

        Args:
            title: Section title
            content: Section content
            level: Heading level
            section_type: Type of section
        """
        section = ReportSection(
            title=title,
            content=content,
            level=level,
            section_type=section_type
        )
        self.sections.append(section)

    def generate_summary(self, aggregated_data: Dict[str, Any]) -> str:
        """
        Generate summary text from aggregated data.
        Override to customize summary generation.

        Args:
            aggregated_data: Aggregated data

        Returns:
            Summary text
        """
        lines = []
        for key, value in aggregated_data.items():
            if isinstance(value, (int, float)):
                lines.append(f"- **{key.replace('_', ' ').title()}**: {value:,.2f}")
            else:
                lines.append(f"- **{key.replace('_', ' ').title()}**: {value}")

        return "\n".join(lines)

    def render_markdown(self) -> str:
        """
        Render report as Markdown.

        Returns:
            Markdown string
        """
        lines = []

        # Header
        lines.append(f"# {self.config.name} Report")
        lines.append("")
        lines.append(f"**Generated:** {DeterministicClock.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append("")

        # Sections
        for section in self.sections:
            # Add heading
            heading_prefix = "#" * section.level
            lines.append(f"{heading_prefix} {section.title}")
            lines.append("")

            # Add content based on type
            if section.section_type == "text":
                lines.append(str(section.content))
            elif section.section_type == "table":
                lines.append(self._render_table_markdown(section.content))
            elif section.section_type == "list":
                if isinstance(section.content, list):
                    for item in section.content:
                        lines.append(f"- {item}")
                else:
                    lines.append(str(section.content))
            else:
                lines.append(str(section.content))

            lines.append("")

        return "\n".join(lines)

    def render_html(self) -> str:
        """
        Render report as HTML.

        Returns:
            HTML string
        """
        html_lines = []

        # HTML header
        html_lines.append("<!DOCTYPE html>")
        html_lines.append("<html>")
        html_lines.append("<head>")
        html_lines.append(f"<title>{self.config.name} Report</title>")
        html_lines.append("<style>")
        html_lines.append("body { font-family: Arial, sans-serif; margin: 40px; background: #f5f5f5; }")
        html_lines.append(".container { max-width: 1200px; margin: 0 auto; background: white; padding: 30px; }")
        html_lines.append("h1 { color: #2c3e50; border-bottom: 3px solid #3498db; }")
        html_lines.append("h2 { color: #34495e; border-bottom: 1px solid #ecf0f1; }")
        html_lines.append("table { border-collapse: collapse; width: 100%; margin: 20px 0; }")
        html_lines.append("th, td { border: 1px solid #ddd; padding: 12px; text-align: left; }")
        html_lines.append("th { background-color: #3498db; color: white; }")
        html_lines.append("</style>")
        html_lines.append("</head>")
        html_lines.append("<body>")
        html_lines.append("<div class='container'>")

        # Header
        html_lines.append(f"<h1>{self.config.name} Report</h1>")
        html_lines.append(f"<p><strong>Generated:</strong> {DeterministicClock.now().strftime('%Y-%m-%d %H:%M:%S')}</p>")

        # Sections
        for section in self.sections:
            heading_tag = f"h{section.level}"
            html_lines.append(f"<{heading_tag}>{section.title}</{heading_tag}>")

            if section.section_type == "text":
                html_lines.append(f"<p>{section.content}</p>")
            elif section.section_type == "table":
                html_lines.append(self._render_table_html(section.content))
            elif section.section_type == "list":
                html_lines.append("<ul>")
                if isinstance(section.content, list):
                    for item in section.content:
                        html_lines.append(f"<li>{item}</li>")
                html_lines.append("</ul>")
            else:
                html_lines.append(f"<pre>{section.content}</pre>")

        html_lines.append("</div>")
        html_lines.append("</body>")
        html_lines.append("</html>")

        return "\n".join(html_lines)

    def render_json(self) -> str:
        """
        Render report as JSON.

        Returns:
            JSON string
        """
        import json

        report_dict = {
            "report_name": self.config.name,
            "generated_at": DeterministicClock.now().isoformat(),
            "sections": [section.dict() for section in self.sections],
            "metadata": self.metadata
        }

        return json.dumps(report_dict, indent=2, default=str)

    def render_excel(self, output_path: str):
        """
        Render report as Excel file.

        Args:
            output_path: Path to save Excel file

        Raises:
            ImportError: If openpyxl not available
        """
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("openpyxl required for Excel export. Install with: pip install openpyxl")

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Report"

        row = 1

        # Header
        sheet.cell(row=row, column=1, value=f"{self.config.name} Report")
        row += 1
        sheet.cell(row=row, column=1, value=f"Generated: {DeterministicClock.now().strftime('%Y-%m-%d %H:%M:%S')}")
        row += 2

        # Sections
        for section in self.sections:
            # Section title
            sheet.cell(row=row, column=1, value=section.title)
            row += 1

            # Section content
            if section.section_type == "table" and isinstance(section.content, list):
                # Write table
                if section.content:
                    # Headers
                    headers = list(section.content[0].keys())
                    for col, header in enumerate(headers, start=1):
                        sheet.cell(row=row, column=col, value=header)
                    row += 1

                    # Data rows
                    for record in section.content:
                        for col, header in enumerate(headers, start=1):
                            sheet.cell(row=row, column=col, value=record.get(header))
                        row += 1
            else:
                sheet.cell(row=row, column=1, value=str(section.content))
                row += 1

            row += 1  # Blank line between sections

        workbook.save(output_path)
        logger.info(f"Excel report saved to {output_path}")

    def _render_table_markdown(self, data: Any) -> str:
        """Render data as Markdown table."""
        if not isinstance(data, list) or not data:
            return str(data)

        # Get headers from first item
        if isinstance(data[0], dict):
            headers = list(data[0].keys())
        else:
            return str(data)

        # Build table
        lines = []

        # Header row
        lines.append("| " + " | ".join(headers) + " |")
        lines.append("| " + " | ".join(["---"] * len(headers)) + " |")

        # Data rows
        for row in data:
            values = [str(row.get(h, "")) for h in headers]
            lines.append("| " + " | ".join(values) + " |")

        return "\n".join(lines)

    def _render_table_html(self, data: Any) -> str:
        """Render data as HTML table."""
        if not isinstance(data, list) or not data:
            return str(data)

        if isinstance(data[0], dict):
            headers = list(data[0].keys())
        else:
            return str(data)

        lines = ["<table>"]

        # Header
        lines.append("<tr>")
        for h in headers:
            lines.append(f"<th>{h}</th>")
        lines.append("</tr>")

        # Rows
        for row in data:
            lines.append("<tr>")
            for h in headers:
                lines.append(f"<td>{row.get(h, '')}</td>")
            lines.append("</tr>")

        lines.append("</table>")

        return "\n".join(lines)

    def execute(self, input_data: Dict[str, Any]) -> AgentResult:
        """
        Execute reporting with aggregation and formatting.

        Args:
            input_data: Must contain data to report on

        Returns:
            AgentResult with rendered report
        """
        # Clear previous sections
        self.sections = []

        # Aggregate data
        logger.info(f"Aggregating data for {self.config.name}")
        aggregated_data = self.aggregate_data(input_data)

        # Add summary if enabled
        if self.config.include_summary:
            summary = self.generate_summary(aggregated_data)
            self.add_section("Summary", summary, level=2, section_type="text")

        # Build sections
        if self.config.include_details:
            logger.info(f"Building report sections")
            sections = self.build_sections(aggregated_data)
            self.sections.extend(sections)

        # Render in requested format
        logger.info(f"Rendering report in {self.config.output_format} format")

        if self.config.output_format == "markdown":
            report_content = self.render_markdown()
        elif self.config.output_format == "html":
            report_content = self.render_html()
        elif self.config.output_format == "json":
            report_content = self.render_json()
        elif self.config.output_format == "excel":
            # For Excel, we don't return content, just save
            output_path = input_data.get("output_path", "report.xlsx")
            self.render_excel(output_path)
            report_content = f"Excel report saved to {output_path}"
        else:
            raise ValueError(f"Unsupported output format: {self.config.output_format}")

        return AgentResult(
            success=True,
            data={
                "report": report_content,
                "format": self.config.output_format,
                "sections_count": len(self.sections)
            },
            metadata={
                "aggregated_data": aggregated_data,
                "generated_at": DeterministicClock.now().isoformat()
            }
        )

    def validate_input(self, input_data: Dict[str, Any]) -> bool:
        """Validate that input contains data to report on."""
        if not input_data:
            self.logger.error("Input data is empty")
            return False

        return True
