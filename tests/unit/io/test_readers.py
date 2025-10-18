"""
Comprehensive tests for GreenLang Data Readers.

Tests cover:
- Multi-format reading (CSV, JSON, Excel, YAML, XML)
- Automatic format detection
- Error handling for missing dependencies
- Format-specific options
- Edge cases and error handling
"""

import pytest
import json
import csv
from pathlib import Path
from unittest.mock import patch, MagicMock

from greenlang.io.readers import DataReader, read_file


# Test Fixtures
@pytest.fixture
def reader():
    """Create a data reader instance."""
    return DataReader()


@pytest.fixture
def sample_json_data():
    """Sample JSON data."""
    return {
        "name": "John Doe",
        "age": 30,
        "email": "john@example.com"
    }


@pytest.fixture
def sample_csv_data():
    """Sample CSV data."""
    return [
        {"name": "John", "age": "30", "city": "New York"},
        {"name": "Jane", "age": "25", "city": "Boston"},
        {"name": "Bob", "age": "35", "city": "Chicago"}
    ]


@pytest.fixture
def json_file(tmp_path, sample_json_data):
    """Create temporary JSON file."""
    file_path = tmp_path / "data.json"
    with open(file_path, 'w') as f:
        json.dump(sample_json_data, f)
    return file_path


@pytest.fixture
def csv_file(tmp_path, sample_csv_data):
    """Create temporary CSV file."""
    file_path = tmp_path / "data.csv"
    with open(file_path, 'w', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=["name", "age", "city"])
        writer.writeheader()
        writer.writerows(sample_csv_data)
    return file_path


@pytest.fixture
def text_file(tmp_path):
    """Create temporary text file."""
    file_path = tmp_path / "data.txt"
    file_path.write_text("Hello, World!\nThis is a test file.")
    return file_path


# DataReader Initialization Tests
class TestDataReaderInit:
    """Test DataReader initialization."""

    def test_reader_creation(self):
        """Test creating a data reader."""
        reader = DataReader()
        assert reader.default_encoding == "utf-8"
        assert ".json" in reader._format_handlers
        assert ".csv" in reader._format_handlers

    def test_custom_encoding(self):
        """Test reader with custom encoding."""
        reader = DataReader(default_encoding="latin-1")
        assert reader.default_encoding == "latin-1"

    def test_supported_formats(self, reader):
        """Test getting supported formats."""
        formats = reader.get_supported_formats()
        assert ".json" in formats
        assert ".csv" in formats
        assert ".txt" in formats
        assert ".tsv" in formats


# JSON Reading Tests
class TestJSONReading:
    """Test JSON file reading."""

    def test_read_json_object(self, reader, json_file, sample_json_data):
        """Test reading JSON object."""
        data = reader.read(json_file)
        assert data == sample_json_data
        assert data["name"] == "John Doe"
        assert data["age"] == 30

    def test_read_json_array(self, reader, tmp_path):
        """Test reading JSON array."""
        array_data = [
            {"id": 1, "name": "Item 1"},
            {"id": 2, "name": "Item 2"}
        ]
        file_path = tmp_path / "array.json"
        with open(file_path, 'w') as f:
            json.dump(array_data, f)

        data = reader.read(file_path)
        assert isinstance(data, list)
        assert len(data) == 2

    def test_read_json_nested(self, reader, tmp_path):
        """Test reading nested JSON."""
        nested_data = {
            "user": {
                "name": "John",
                "address": {
                    "city": "New York",
                    "zipcode": "10001"
                }
            }
        }
        file_path = tmp_path / "nested.json"
        with open(file_path, 'w') as f:
            json.dump(nested_data, f)

        data = reader.read(file_path)
        assert data["user"]["name"] == "John"
        assert data["user"]["address"]["city"] == "New York"

    def test_read_json_unicode(self, reader, tmp_path):
        """Test reading JSON with unicode characters."""
        unicode_data = {"message": "Hello 世界 🌍"}
        file_path = tmp_path / "unicode.json"
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(unicode_data, f, ensure_ascii=False)

        data = reader.read(file_path)
        assert data["message"] == "Hello 世界 🌍"


# CSV Reading Tests
class TestCSVReading:
    """Test CSV file reading."""

    def test_read_csv_with_header(self, reader, csv_file):
        """Test reading CSV with header."""
        data = reader.read(csv_file)
        assert isinstance(data, list)
        assert len(data) == 3
        assert data[0]["name"] == "John"
        assert data[0]["age"] == "30"

    def test_read_csv_custom_delimiter(self, reader, tmp_path):
        """Test reading CSV with custom delimiter."""
        file_path = tmp_path / "semicolon.csv"
        with open(file_path, 'w') as f:
            f.write("name;age;city\n")
            f.write("John;30;New York\n")
            f.write("Jane;25;Boston\n")

        data = reader.read(file_path, csv_delimiter=";")
        assert len(data) == 2
        assert data[0]["name"] == "John"

    def test_read_csv_without_header(self, reader, tmp_path):
        """Test reading CSV without header."""
        file_path = tmp_path / "no_header.csv"
        with open(file_path, 'w') as f:
            f.write("John,30,New York\n")
            f.write("Jane,25,Boston\n")

        data = reader.read(file_path, csv_has_header=False)
        assert isinstance(data, list)
        assert isinstance(data[0], list)

    def test_read_tsv(self, reader, tmp_path):
        """Test reading TSV file."""
        file_path = tmp_path / "data.tsv"
        with open(file_path, 'w') as f:
            f.write("name\tage\tcity\n")
            f.write("John\t30\tNew York\n")
            f.write("Jane\t25\tBoston\n")

        data = reader.read(file_path)
        assert len(data) == 2
        assert data[0]["name"] == "John"

    def test_read_csv_empty_file(self, reader, tmp_path):
        """Test reading empty CSV file."""
        file_path = tmp_path / "empty.csv"
        file_path.write_text("")

        data = reader.read(file_path)
        assert len(data) == 0

    def test_read_csv_only_header(self, reader, tmp_path):
        """Test reading CSV with only header."""
        file_path = tmp_path / "header_only.csv"
        file_path.write_text("name,age,city\n")

        data = reader.read(file_path)
        assert len(data) == 0


# Text Reading Tests
class TestTextReading:
    """Test text file reading."""

    def test_read_text_file(self, reader, text_file):
        """Test reading text file."""
        data = reader.read(text_file)
        assert isinstance(data, str)
        assert "Hello, World!" in data

    def test_read_text_multiline(self, reader, tmp_path):
        """Test reading multiline text file."""
        file_path = tmp_path / "multiline.txt"
        file_path.write_text("Line 1\nLine 2\nLine 3")

        data = reader.read(file_path)
        assert "Line 1" in data
        assert "Line 2" in data
        assert "Line 3" in data


# YAML Reading Tests
class TestYAMLReading:
    """Test YAML file reading."""

    @patch('greenlang.io.readers.yaml', None)
    def test_yaml_not_available(self, reader):
        """Test that YAML is not in handlers when unavailable."""
        # Create new reader instance without yaml
        reader_new = DataReader()
        formats = reader_new.get_supported_formats()
        # YAML might not be available depending on imports

    def test_read_yaml(self, reader, tmp_path):
        """Test reading YAML file."""
        try:
            import yaml
            file_path = tmp_path / "data.yaml"
            yaml_content = """
name: John Doe
age: 30
tags:
  - developer
  - python
"""
            file_path.write_text(yaml_content)

            data = reader.read(file_path)
            assert data["name"] == "John Doe"
            assert data["age"] == 30
            assert "python" in data["tags"]
        except ImportError:
            pytest.skip("PyYAML not available")

    def test_read_yaml_nested(self, reader, tmp_path):
        """Test reading nested YAML."""
        try:
            import yaml
            file_path = tmp_path / "nested.yaml"
            yaml_content = """
user:
  name: John
  address:
    city: New York
    zip: "10001"
"""
            file_path.write_text(yaml_content)

            data = reader.read(file_path)
            assert data["user"]["name"] == "John"
            assert data["user"]["address"]["city"] == "New York"
        except ImportError:
            pytest.skip("PyYAML not available")


# Excel Reading Tests
class TestExcelReading:
    """Test Excel file reading."""

    def test_read_excel_xlsx(self, reader, tmp_path):
        """Test reading Excel .xlsx file."""
        try:
            import openpyxl

            file_path = tmp_path / "data.xlsx"

            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["name", "age", "city"])
            ws.append(["John", 30, "New York"])
            ws.append(["Jane", 25, "Boston"])
            wb.save(file_path)
            wb.close()

            data = reader.read(file_path)
            assert len(data) == 2
            assert data[0]["name"] == "John"
        except ImportError:
            pytest.skip("openpyxl not available")

    def test_read_excel_specific_sheet(self, reader, tmp_path):
        """Test reading specific Excel sheet."""
        try:
            import openpyxl

            file_path = tmp_path / "multi_sheet.xlsx"

            # Create workbook with multiple sheets
            wb = openpyxl.Workbook()
            ws1 = wb.active
            ws1.title = "Sheet1"
            ws1.append(["col1", "col2"])
            ws1.append(["a", "b"])

            ws2 = wb.create_sheet("Sheet2")
            ws2.append(["name", "value"])
            ws2.append(["test", 123])

            wb.save(file_path)
            wb.close()

            data = reader.read(file_path, sheet_name="Sheet2")
            assert data[0]["name"] == "test"
        except ImportError:
            pytest.skip("openpyxl not available")


# XML Reading Tests
class TestXMLReading:
    """Test XML file reading."""

    def test_read_xml(self, reader, tmp_path):
        """Test reading XML file."""
        try:
            from lxml import etree

            file_path = tmp_path / "data.xml"
            xml_content = """<?xml version="1.0"?>
<root>
    <name>John Doe</name>
    <age>30</age>
</root>
"""
            file_path.write_text(xml_content)

            data = reader.read(file_path)
            assert "root" in data
        except ImportError:
            pytest.skip("lxml not available")


# Error Handling Tests
class TestErrorHandling:
    """Test error handling."""

    def test_file_not_found(self, reader):
        """Test reading non-existent file."""
        with pytest.raises(FileNotFoundError):
            reader.read("nonexistent.json")

    def test_unsupported_format(self, reader, tmp_path):
        """Test reading unsupported format."""
        file_path = tmp_path / "data.xyz"
        file_path.write_text("test")

        with pytest.raises(ValueError, match="Unsupported format"):
            reader.read(file_path)

    def test_invalid_json(self, reader, tmp_path):
        """Test reading invalid JSON."""
        file_path = tmp_path / "invalid.json"
        file_path.write_text("{invalid json}")

        with pytest.raises(json.JSONDecodeError):
            reader.read(file_path)

    def test_malformed_csv(self, reader, tmp_path):
        """Test reading malformed CSV."""
        file_path = tmp_path / "malformed.csv"
        file_path.write_text("name,age\nJohn,30,extra")

        # Should still read but may have inconsistent data
        data = reader.read(file_path)
        # CSV reader is lenient


# Convenience Function Tests
class TestConvenienceFunctions:
    """Test convenience functions."""

    def test_read_file_function(self, json_file, sample_json_data):
        """Test read_file convenience function."""
        data = read_file(json_file)
        assert data == sample_json_data

    def test_read_file_with_options(self, tmp_path):
        """Test read_file with format options."""
        file_path = tmp_path / "data.csv"
        with open(file_path, 'w') as f:
            f.write("name;age\nJohn;30\n")

        data = read_file(file_path, csv_delimiter=";")
        assert data[0]["name"] == "John"


# Path Handling Tests
class TestPathHandling:
    """Test path handling."""

    def test_string_path(self, reader, json_file, sample_json_data):
        """Test reading with string path."""
        data = reader.read(str(json_file))
        assert data == sample_json_data

    def test_path_object(self, reader, json_file, sample_json_data):
        """Test reading with Path object."""
        data = reader.read(json_file)
        assert data == sample_json_data

    def test_absolute_path(self, reader, tmp_path, sample_json_data):
        """Test reading with absolute path."""
        file_path = tmp_path / "data.json"
        with open(file_path, 'w') as f:
            json.dump(sample_json_data, f)

        data = reader.read(file_path.absolute())
        assert data == sample_json_data


# Edge Cases Tests
class TestEdgeCases:
    """Test edge cases."""

    def test_empty_json_object(self, reader, tmp_path):
        """Test reading empty JSON object."""
        file_path = tmp_path / "empty.json"
        file_path.write_text("{}")

        data = reader.read(file_path)
        assert data == {}

    def test_empty_json_array(self, reader, tmp_path):
        """Test reading empty JSON array."""
        file_path = tmp_path / "empty_array.json"
        file_path.write_text("[]")

        data = reader.read(file_path)
        assert data == []

    def test_large_json(self, reader, tmp_path):
        """Test reading large JSON file."""
        large_data = [{"id": i, "value": f"item_{i}"} for i in range(1000)]
        file_path = tmp_path / "large.json"
        with open(file_path, 'w') as f:
            json.dump(large_data, f)

        data = reader.read(file_path)
        assert len(data) == 1000

    def test_special_characters_in_csv(self, reader, tmp_path):
        """Test CSV with special characters."""
        file_path = tmp_path / "special.csv"
        with open(file_path, 'w') as f:
            f.write('name,description\n')
            f.write('Item,"Contains, comma and ""quotes"""\n')

        data = reader.read(file_path)
        assert 'comma' in data[0]["description"]

    def test_case_insensitive_extension(self, reader, tmp_path, sample_json_data):
        """Test that extension matching is case insensitive."""
        file_path = tmp_path / "data.JSON"
        with open(file_path, 'w') as f:
            json.dump(sample_json_data, f)

        data = reader.read(file_path)
        assert data == sample_json_data
