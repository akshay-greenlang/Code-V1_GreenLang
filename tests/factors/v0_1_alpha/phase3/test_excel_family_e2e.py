# -*- coding: utf-8 -*-
"""Phase 3 / Wave 2.0 — Excel-family end-to-end tests (parametrized).

Drives the unified :class:`IngestionPipelineRunner` against five
Wave 2.0 Excel-family fixtures (EPA, eGRID, CEA, BEE, IEA). For every
source this module asserts:

  1. **Snapshot match** — :meth:`Phase3<...>ExcelParser.parse_bytes`
     output is byte-stable against the committed golden snapshot.
  2. **Pipeline ladder advance** — the unified runner walks the run
     status from ``created -> fetched -> parsed -> normalized ->
     validated -> deduped -> staged`` end-to-end without manual stage
     calls.
  3. **Family-specific validation rejection** — each adapter rejects a
     deliberately broken fixture with :class:`ParserDispatchError`. The
     four broken-fixture shapes exercised are: changed tab name,
     missing required column, wrong unit string, vintage mismatch.

The Phase 2 + Phase 3 regression remain GREEN; Wave 2.0 only adds new
tests, never modifies an existing one.

References
----------
- ``docs/factors/PHASE_3_PLAN.md`` §"Wave 2.0 — Excel parsers"
- ``greenlang/factors/ingestion/parsers/_phase3_adapters.py`` — Wave 2.0
  adapter classes.
- ``tests/factors/v0_1_alpha/phase3/test_defra_reference_e2e.py`` —
  Wave 1.5 reference template.
"""
from __future__ import annotations

import io
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable, Dict, List, Tuple

import pytest

from greenlang.factors.ingestion.exceptions import (
    IngestionError,
    ParserDispatchError,
)
from greenlang.factors.ingestion.parsers._phase3_adapters import (
    PHASE3_BEE_PARSER_VERSION,
    PHASE3_BEE_SOURCE_ID,
    PHASE3_BEE_SOURCE_URN,
    PHASE3_CEA_PARSER_VERSION,
    PHASE3_CEA_SOURCE_ID,
    PHASE3_CEA_SOURCE_URN,
    PHASE3_EGRID_PARSER_VERSION,
    PHASE3_EGRID_SOURCE_ID,
    PHASE3_EGRID_SOURCE_URN,
    PHASE3_EPA_PARSER_VERSION,
    PHASE3_EPA_SOURCE_ID,
    PHASE3_EPA_SOURCE_URN,
    PHASE3_IEA_PARSER_VERSION,
    PHASE3_IEA_SOURCE_ID,
    PHASE3_IEA_SOURCE_URN,
    Phase3BEEExcelParser,
    Phase3CEAExcelParser,
    Phase3EGridExcelParser,
    Phase3EPAExcelParser,
    Phase3IEAExcelParser,
)
from greenlang.factors.ingestion.pipeline import RunStatus
from tests.factors.v0_1_alpha.phase3.fixtures import (
    _build_bee_fixture,
    _build_cea_fixture,
    _build_egrid_fixture,
    _build_epa_fixture,
    _build_iea_fixture,
)
from tests.factors.v0_1_alpha.phase3.parser_snapshots._helper import (
    compare_to_snapshot,
    regenerate_if_env,
)


# ---------------------------------------------------------------------------
# Per-source dispatch table — drives every parametrized test below.
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class ExcelSourceCase:
    """Wiring for one Wave 2.0 Excel-family source."""

    source_id: str
    source_urn: str
    parser_version: str
    parser_cls: Any
    fixture_filename: str
    fixture_module: Any
    snapshot_id: str
    expected_row_count: int
    source_version: str = "2024.1"


_EPA_CASE = ExcelSourceCase(
    source_id=PHASE3_EPA_SOURCE_ID,
    source_urn=PHASE3_EPA_SOURCE_URN,
    parser_version=PHASE3_EPA_PARSER_VERSION,
    parser_cls=Phase3EPAExcelParser,
    fixture_filename="epa_2024_mini.xlsx",
    fixture_module=_build_epa_fixture,
    snapshot_id="epa_2024",
    expected_row_count=5 + 3,  # stationary + mobile
)

_EGRID_CASE = ExcelSourceCase(
    source_id=PHASE3_EGRID_SOURCE_ID,
    source_urn=PHASE3_EGRID_SOURCE_URN,
    parser_version=PHASE3_EGRID_PARSER_VERSION,
    parser_cls=Phase3EGridExcelParser,
    fixture_filename="egrid_2024_mini.xlsx",
    fixture_module=_build_egrid_fixture,
    snapshot_id="egrid_2024",
    expected_row_count=5 + 3,
)

_CEA_CASE = ExcelSourceCase(
    source_id=PHASE3_CEA_SOURCE_ID,
    source_urn=PHASE3_CEA_SOURCE_URN,
    parser_version=PHASE3_CEA_PARSER_VERSION,
    parser_cls=Phase3CEAExcelParser,
    fixture_filename="cea_2024_mini.xlsx",
    fixture_module=_build_cea_fixture,
    snapshot_id="cea_2024",
    expected_row_count=4,
    source_version="2023-24",
)

_BEE_CASE = ExcelSourceCase(
    source_id=PHASE3_BEE_SOURCE_ID,
    source_urn=PHASE3_BEE_SOURCE_URN,
    parser_version=PHASE3_BEE_PARSER_VERSION,
    parser_cls=Phase3BEEExcelParser,
    fixture_filename="bee_2024_mini.xlsx",
    fixture_module=_build_bee_fixture,
    snapshot_id="bee_2024",
    expected_row_count=5,
)

_IEA_CASE = ExcelSourceCase(
    source_id=PHASE3_IEA_SOURCE_ID,
    source_urn=PHASE3_IEA_SOURCE_URN,
    parser_version=PHASE3_IEA_PARSER_VERSION,
    parser_cls=Phase3IEAExcelParser,
    fixture_filename="iea_2024_mini.xlsx",
    fixture_module=_build_iea_fixture,
    snapshot_id="iea_2024",
    expected_row_count=5,
)


_ALL_CASES: Tuple[ExcelSourceCase, ...] = (
    _EPA_CASE,
    _EGRID_CASE,
    _CEA_CASE,
    _BEE_CASE,
    _IEA_CASE,
)


def _materialise_fixture(case: ExcelSourceCase) -> Path:
    """Idempotently materialise the deterministic fixture for ``case``."""
    fixture_dir = Path(__file__).resolve().parent / "fixtures"
    target = fixture_dir / case.fixture_filename
    case.fixture_module.ensure_fixture(target)
    return target


# ---------------------------------------------------------------------------
# 1. Snapshot tests — one per source.
# ---------------------------------------------------------------------------


@pytest.mark.parametrize(
    "case",
    _ALL_CASES,
    ids=[c.source_id for c in _ALL_CASES],
)
def test_excel_family_parser_snapshot_matches_golden(
    case: ExcelSourceCase,
) -> None:
    """Each Wave 2.0 parser's output is byte-stable against its golden."""
    fixture_path = _materialise_fixture(case)
    parser = case.parser_cls()
    raw_bytes = fixture_path.read_bytes()
    rows = parser.parse_bytes(
        raw_bytes,
        artifact_uri=f"file://{case.fixture_filename}",
        artifact_sha256="0" * 64,
    )
    # Strip volatile fields (wall-clock timestamps) before snapshot
    # comparison so the golden does not encode time-of-run.
    for row in rows:
        row.pop("published_at", None)
        ext = row.get("extraction") or {}
        if isinstance(ext, dict):
            ext.pop("ingested_at", None)
        review = row.get("review") or {}
        if isinstance(review, dict):
            for vol_key in ("reviewed_at", "approved_at"):
                review.pop(vol_key, None)

    assert len(rows) == case.expected_row_count, (
        f"{case.source_id} parsed {len(rows)} rows; expected "
        f"{case.expected_row_count}"
    )
    regenerate_if_env(case.snapshot_id, case.parser_version, rows)
    compare_to_snapshot(case.snapshot_id, case.parser_version, rows)


# ---------------------------------------------------------------------------
# 2. Pipeline-ladder e2e — one per source.
# ---------------------------------------------------------------------------


@pytest.mark.parametrize(
    "case",
    _ALL_CASES,
    ids=[c.source_id for c in _ALL_CASES],
)
def test_excel_family_run_advances_through_canonical_status_ladder(
    phase3_runner_raw,
    phase3_run_repo,
    case: ExcelSourceCase,
) -> None:
    """``runner.run()`` walks created -> fetched -> ... -> staged.

    Wave 2.0 mirror of ``test_defra_run_advances_through_canonical_status_ladder``
    — parametrized over the five Excel-family sources.
    """
    fixture_path = _materialise_fixture(case)
    run = phase3_runner_raw.run(
        source_id=case.source_id,
        source_url=str(fixture_path.resolve()),
        source_urn=case.source_urn,
        source_version=case.source_version,
        operator="bot:test-wave2",
        auto_stage=True,
    )

    assert run.status in (RunStatus.STAGED, RunStatus.REVIEW_REQUIRED), (
        f"{case.source_id} unexpected terminal status: {run.status.value}"
    )

    conn = phase3_run_repo._memory_conn  # type: ignore[attr-defined]
    cur = conn.execute(
        "SELECT stage FROM ingestion_run_stage_history "
        "WHERE run_id = ? ORDER BY pk_id ASC",
        (run.run_id,),
    )
    stages_observed = [row[0] for row in cur.fetchall()]
    expected_ladder = [
        "fetch", "parse", "normalize", "validate", "dedupe", "stage",
    ]
    idx = 0
    for stage in stages_observed:
        if idx < len(expected_ladder) and stage == expected_ladder[idx]:
            idx += 1
    assert idx == len(expected_ladder), (
        f"{case.source_id} canonical ladder not observed; "
        f"saw {stages_observed!r}"
    )


# ---------------------------------------------------------------------------
# 3. Family-specific validation — negative-path parametrize.
# ---------------------------------------------------------------------------


def _build_broken_workbook(
    case: ExcelSourceCase,
    *,
    break_kind: str,
) -> bytes:
    """Build a deterministic broken-shape workbook for ``case``.

    ``break_kind`` is one of:
      * ``"tab_name"``   — rename a required tab.
      * ``"column"``     — drop a required column from the header row.
      * ``"unit"``       — write a unit string outside the registry pin.
      * ``"vintage"``    — write a vintage label outside the registry window.

    The function calls into the source's fixture module to materialise
    the canonical workbook bytes, then unpacks + mutates + repacks via
    :mod:`openpyxl` so the rest of the workbook (creation stamps, sheet
    order) remains byte-deterministic.
    """
    import openpyxl  # noqa: PLC0415

    canonical = case.fixture_module.build_workbook_bytes()
    wb = openpyxl.load_workbook(io.BytesIO(canonical), data_only=False)

    def _first_sheet() -> Any:
        return wb[wb.sheetnames[0]]

    if break_kind == "tab_name":
        # Rename the first required tab to a sentinel that's NOT in the
        # adapter's required_tabs set.
        ws = _first_sheet()
        ws.title = "BROKEN_TAB_NAME"
    elif break_kind == "column":
        # Drop the second column (the unit column for every Wave 2.0
        # source); the adapter's required_headers check rejects this.
        ws = _first_sheet()
        # Overwrite cell B1 with an unrelated label so the required
        # column 'unit' is no longer present in the header row.
        ws.cell(row=1, column=2).value = "totally_unrelated_column"
    elif break_kind == "unit":
        # Replace the unit value on the first data row with a unit
        # outside the allowed set.
        ws = _first_sheet()
        ws.cell(row=2, column=2).value = "wrong_unit_xyz"
    elif break_kind == "vintage":
        # Replace the vintage column on the first data row with a value
        # outside the registry vintage window. Skip cleanly if this
        # source has no vintage column.
        if not getattr(case.parser_cls, "vintage_column", None):
            pytest.skip(
                f"{case.source_id} has no vintage column; "
                "vintage drift is enforced via constructor pin only"
            )
        ws = _first_sheet()
        # Find the vintage column index by scanning the header row.
        header_cells = [c.value for c in ws[1]]
        try:
            col_idx = header_cells.index(case.parser_cls.vintage_column) + 1
        except ValueError:
            pytest.skip(
                f"{case.source_id} fixture is missing the vintage column"
            )
        ws.cell(row=2, column=col_idx).value = "9999.9"
    else:  # pragma: no cover — guarded by parametrize id list
        raise AssertionError(f"unknown break_kind: {break_kind!r}")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# Build the parametrize matrix as (case, break_kind) pairs so pytest's
# test id reads like ``epa_hub-tab_name``.
_BREAK_KINDS = ("tab_name", "column", "unit", "vintage")
_NEGATIVE_MATRIX = [
    (c, b) for c in _ALL_CASES for b in _BREAK_KINDS
]
_NEGATIVE_IDS = [
    f"{c.source_id}-{b}" for c, b in _NEGATIVE_MATRIX
]


@pytest.mark.parametrize(
    ("case", "break_kind"),
    _NEGATIVE_MATRIX,
    ids=_NEGATIVE_IDS,
)
def test_excel_family_validation_rejects_broken_fixture(
    case: ExcelSourceCase,
    break_kind: str,
) -> None:
    """Each adapter raises :class:`ParserDispatchError` on broken shapes.

    Verifies the family-specific validation gates (tab name, column
    presence, unit pin, vintage pin) reject the correct kind of failure.
    """
    raw_bytes = _build_broken_workbook(case, break_kind=break_kind)
    parser = case.parser_cls(source_version=case.source_version)
    with pytest.raises((ParserDispatchError, IngestionError)) as exc_info:
        parser.parse_bytes(
            raw_bytes,
            artifact_uri=f"file://broken-{break_kind}.xlsx",
            artifact_sha256="0" * 64,
        )
    msg = str(exc_info.value).lower()
    # Sanity-check the error message names the violation. We deliberately
    # avoid asserting on a single phrase so the message wording can drift
    # without a test rewrite.
    if break_kind == "tab_name":
        assert "tab" in msg or "missing" in msg
    elif break_kind == "column":
        assert "column" in msg or "header" in msg or "missing" in msg
    elif break_kind == "unit":
        assert "unit" in msg or "registry" in msg
    elif break_kind == "vintage":
        assert "vintage" in msg or "window" in msg


# ---------------------------------------------------------------------------
# 4. Registry dispatch — Phase3 registry carries every Wave 2.0 source.
# ---------------------------------------------------------------------------


def test_phase3_registry_carries_every_wave2_source() -> None:
    """The :func:`build_phase3_registry` registry has all 5 Wave 2.0 sources."""
    from greenlang.factors.ingestion.parsers._phase3_adapters import (
        build_phase3_registry,
    )

    registry = build_phase3_registry()
    for case in _ALL_CASES:
        parser = registry.get(case.source_id)
        assert parser is not None, (
            f"Wave 2.0 source {case.source_id!r} not in Phase3 registry"
        )
        assert parser.parser_version == case.parser_version, (
            f"{case.source_id} parser_version drift: "
            f"registry={parser.parser_version!r} "
            f"expected={case.parser_version!r}"
        )
