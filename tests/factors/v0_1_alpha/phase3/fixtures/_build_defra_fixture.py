# -*- coding: utf-8 -*-
"""Build the deterministic synthetic DEFRA / DESNZ Excel fixture.

The Phase 3 Wave 1.5 plan calls for a byte-identical, network-free,
DEFRA-shaped Excel artifact that the unified ingestion pipeline can
fetch -> parse -> normalize -> validate -> dedupe -> stage -> publish
end-to-end.

Determinism contract
--------------------
- Every row is written from a frozen, sorted-key tuple of cell values
  so iteration order is invariant across Python versions.
- ``openpyxl`` is invoked with ``Workbook()`` defaults; we explicitly
  pin ``wb.properties.creator`` / ``created`` / ``modified`` to a fixed
  ISO-8601 timestamp so the embedded ``xl/core.xml`` does not embed a
  wall-clock value.
- The file is written to a ``BytesIO`` buffer and the bytes flushed
  with ``Path.write_bytes`` so we never depend on filesystem mtime.

References
----------
- ``docs/factors/PHASE_3_PLAN.md`` §"Reference source: DEFRA Excel
  end-to-end"
- ``docs/factors/PHASE_3_EXIT_CHECKLIST.md`` Block 3 ("Excel-family
  validation")
"""
from __future__ import annotations

import io
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, List, Tuple

#: Deterministic stamp written into the workbook's core properties so the
#: embedded XML does NOT carry a wall-clock value.
_FROZEN_STAMP = datetime(2026, 4, 28, 0, 0, 0, tzinfo=timezone.utc)

#: Sheet 1 — Stationary Combustion. Five rows of plausible DEFRA shapes.
#: Column order intentionally matches the DESNZ parser's ``scope1_fuels``
#: section (``fuel_type, unit, co2_factor, ch4_factor, n2o_factor, notes``).
STATIONARY_HEADERS: Tuple[str, ...] = (
    "fuel_type",
    "unit",
    "co2_factor",
    "ch4_factor",
    "n2o_factor",
    "notes",
)
STATIONARY_ROWS: Tuple[Tuple[Any, ...], ...] = (
    ("natural_gas", "kwh", 0.18293, 0.00033, 0.00010, "DEFRA 2025 default"),
    ("diesel", "litre", 2.51233, 0.00148, 0.01441, "DEFRA 2025 default"),
    ("lpg", "kwh", 0.21449, 0.00007, 0.00009, "DEFRA 2025 default"),
    ("coal_industrial", "kg", 2.41062, 0.00012, 0.00031, "DEFRA 2025 default"),
    ("biomass_wood_pellets", "kwh", 0.01540, 0.00031, 0.00005, "Biogenic component reported separately"),
)

#: Sheet 2 — Fuel Conversion (kWh<->volume helper rows). Same column
#: shape as STATIONARY_HEADERS for parser symmetry; the values differ.
FUEL_CONVERSION_HEADERS: Tuple[str, ...] = STATIONARY_HEADERS
FUEL_CONVERSION_ROWS: Tuple[Tuple[Any, ...], ...] = (
    ("petrol", "litre", 2.16802, 0.00250, 0.00220, "Avg of leaded/unleaded — DEFRA 2025"),
    ("aviation_kerosene", "litre", 2.54453, 0.00010, 0.00080, "DEFRA 2025 default"),
    ("heavy_fuel_oil", "litre", 3.16660, 0.00120, 0.00250, "DEFRA 2025 default"),
    ("burning_oil", "litre", 2.54035, 0.00038, 0.00200, "DEFRA 2025 default"),
    ("compressed_natural_gas", "kg", 2.53718, 0.00037, 0.00007, "DEFRA 2025 default"),
)


def build_workbook_bytes() -> bytes:
    """Return the deterministic xlsx bytes (no I/O).

    A fresh :class:`openpyxl.Workbook` is created per call, populated
    with the frozen sheets, then saved to a :class:`BytesIO` buffer.
    The function is pure: identical inputs (constants above) -> identical
    output bytes across runs.
    """
    import openpyxl  # noqa: PLC0415 — heavyweight import deferred.

    wb = openpyxl.Workbook()
    # Pin creation/modified timestamps so embedded XML is byte-stable.
    wb.properties.creator = "GreenLang Factors / Phase 3 fixture"
    wb.properties.created = _FROZEN_STAMP
    wb.properties.modified = _FROZEN_STAMP
    wb.properties.lastModifiedBy = "phase3-fixture"

    # Sheet 1 — Stationary Combustion (drop the auto "Sheet" Sheet first).
    default = wb.active
    default.title = "Stationary Combustion"
    default.append(list(STATIONARY_HEADERS))
    for row in STATIONARY_ROWS:
        default.append(list(row))

    # Sheet 2 — Fuel Conversion.
    sheet2 = wb.create_sheet("Fuel Conversion")
    sheet2.append(list(FUEL_CONVERSION_HEADERS))
    for row in FUEL_CONVERSION_ROWS:
        sheet2.append(list(row))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def ensure_fixture(path: Path) -> Path:
    """Materialise the DEFRA fixture at ``path`` if it does not exist.

    Idempotent: if the file already exists, returns immediately without
    re-writing (which would be a no-op anyway since the content is byte-
    deterministic). If the file is absent, creates parent directories
    and writes the deterministic bytes.

    Returns:
        ``path`` (for ergonomic chaining at fixture-collection time).
    """
    p = Path(path)
    if p.exists():
        return p
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_bytes(build_workbook_bytes())
    return p


def all_rows() -> List[Tuple[str, Tuple[Any, ...], Tuple[Tuple[Any, ...], ...]]]:
    """Return ``[(sheet_name, headers, rows), ...]`` for assertions.

    The Phase 3 e2e tests use this to count expected rows without
    re-reading the xlsx (which would tie the test to ``openpyxl`` again).
    """
    return [
        ("Stationary Combustion", STATIONARY_HEADERS, STATIONARY_ROWS),
        ("Fuel Conversion", FUEL_CONVERSION_HEADERS, FUEL_CONVERSION_ROWS),
    ]


__all__ = [
    "STATIONARY_HEADERS",
    "STATIONARY_ROWS",
    "FUEL_CONVERSION_HEADERS",
    "FUEL_CONVERSION_ROWS",
    "build_workbook_bytes",
    "ensure_fixture",
    "all_rows",
]


if __name__ == "__main__":  # pragma: no cover — manual regen helper
    fixture_path = Path(__file__).resolve().parent / "defra_2025_mini.xlsx"
    ensure_fixture(fixture_path)
    print(f"wrote {fixture_path} ({fixture_path.stat().st_size} bytes)")
