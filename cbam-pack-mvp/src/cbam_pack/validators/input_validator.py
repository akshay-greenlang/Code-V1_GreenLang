"""
Input Validator

Validates import ledger and config files against schemas and business rules.
Implements fail-fast validation with actionable error messages.
"""

import csv
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Optional

import yaml

from cbam_pack.errors import (
    ErrorLocation,
    MissingRequiredColumnError,
    InvalidDataTypeError,
    InvalidEnumValueError,
    InvalidCNCodeFormatError,
    UnsupportedCNCodeError,
    InvalidCountryCodeError,
    NegativeQuantityError,
    UnknownUnitError,
    DuplicateLineIdError,
    MissingConfigFieldError,
    ValidationError,
)
from cbam_pack.models import (
    ImportLineItem,
    CBAMConfig,
    Quarter,
    Unit,
    VALID_COUNTRY_CODES,
)


@dataclass
class ValidationResult:
    """Result of input validation."""
    is_valid: bool
    errors: list[ValidationError] = field(default_factory=list)
    validated_lines: list[ImportLineItem] = field(default_factory=list)
    validated_config: Optional[CBAMConfig] = None

    @property
    def first_error(self) -> Optional[ValidationError]:
        """Get the first validation error."""
        return self.errors[0] if self.errors else None


class InputValidator:
    """
    Validates CBAM import data and configuration.

    Implements fail-fast validation per PRD requirements.
    """

    REQUIRED_COLUMNS = {
        "line_id",
        "quarter",
        "year",
        "cn_code",
        "product_description",
        "country_of_origin",
        "quantity",
        "unit",
    }

    OPTIONAL_COLUMNS = {
        "supplier_id",
        "installation_id",
        "supplier_direct_emissions",
        "supplier_indirect_emissions",
        "supplier_certificate_ref",
    }

    VALID_QUARTERS = {"Q1", "Q2", "Q3", "Q4"}
    VALID_UNITS = {"kg", "tonnes"}

    def __init__(self, fail_fast: bool = True):
        """
        Initialize validator.

        Args:
            fail_fast: If True, stop at first error. If False, collect all errors.
        """
        self.fail_fast = fail_fast

    def validate_imports(
        self,
        file_path: Path,
    ) -> ValidationResult:
        """
        Validate an import ledger file.

        Args:
            file_path: Path to CSV or XLSX file

        Returns:
            ValidationResult with validated lines or errors
        """
        errors: list[ValidationError] = []
        validated_lines: list[ImportLineItem] = []

        # Read file
        try:
            rows = self._read_file(file_path)
        except Exception as e:
            errors.append(
                ValidationError(
                    f"Failed to read file: {e}",
                    location=ErrorLocation(file=str(file_path)),
                )
            )
            return ValidationResult(is_valid=False, errors=errors)

        if not rows:
            errors.append(
                ValidationError(
                    "Import file is empty",
                    location=ErrorLocation(file=str(file_path)),
                )
            )
            return ValidationResult(is_valid=False, errors=errors)

        # Check required columns
        first_row = rows[0]
        columns = set(first_row.keys())

        missing_columns = self.REQUIRED_COLUMNS - columns
        if missing_columns:
            for col in sorted(missing_columns):
                errors.append(
                    MissingRequiredColumnError(
                        col,
                        location=ErrorLocation(file=str(file_path)),
                    )
                )
                if self.fail_fast:
                    return ValidationResult(is_valid=False, errors=errors)

            if errors:
                return ValidationResult(is_valid=False, errors=errors)

        # Track line_ids for duplicate detection
        seen_line_ids: dict[str, int] = {}

        # Validate each row
        for row_num, row in enumerate(rows, start=2):  # Start at 2 (1 is header)
            result = self._validate_row(
                row=row,
                row_num=row_num,
                file_path=file_path,
                seen_line_ids=seen_line_ids,
            )

            if result.is_valid:
                validated_lines.append(result.validated_lines[0])
            else:
                errors.extend(result.errors)
                if self.fail_fast:
                    return ValidationResult(is_valid=False, errors=errors)

        if errors:
            return ValidationResult(is_valid=False, errors=errors)

        return ValidationResult(
            is_valid=True,
            validated_lines=validated_lines,
        )

    def validate_config(
        self,
        file_path: Path,
    ) -> ValidationResult:
        """
        Validate a CBAM configuration file.

        Args:
            file_path: Path to YAML config file

        Returns:
            ValidationResult with validated config or errors
        """
        errors: list[ValidationError] = []

        # Read YAML
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                data = yaml.safe_load(f)
        except Exception as e:
            errors.append(
                ValidationError(
                    f"Failed to read config file: {e}",
                    location=ErrorLocation(file=str(file_path)),
                )
            )
            return ValidationResult(is_valid=False, errors=errors)

        if not data:
            errors.append(
                ValidationError(
                    "Config file is empty",
                    location=ErrorLocation(file=str(file_path)),
                )
            )
            return ValidationResult(is_valid=False, errors=errors)

        # Validate required fields
        required_fields = [
            "declarant",
            "declarant.name",
            "declarant.eori_number",
            "declarant.address",
            "declarant.address.street",
            "declarant.address.city",
            "declarant.address.postal_code",
            "declarant.address.country",
            "declarant.contact",
            "declarant.contact.name",
            "declarant.contact.email",
            "reporting_period",
            "reporting_period.quarter",
            "reporting_period.year",
        ]

        for field_path in required_fields:
            if not self._has_field(data, field_path):
                errors.append(
                    MissingConfigFieldError(
                        field_path,
                        location=ErrorLocation(file=str(file_path)),
                    )
                )
                if self.fail_fast:
                    return ValidationResult(is_valid=False, errors=errors)

        if errors:
            return ValidationResult(is_valid=False, errors=errors)

        # Try to parse with Pydantic
        try:
            config = CBAMConfig.model_validate(data)
        except Exception as e:
            errors.append(
                ValidationError(
                    f"Config validation failed: {e}",
                    location=ErrorLocation(file=str(file_path)),
                )
            )
            return ValidationResult(is_valid=False, errors=errors)

        return ValidationResult(
            is_valid=True,
            validated_config=config,
        )

    def _read_file(self, file_path: Path) -> list[dict[str, Any]]:
        """Read a CSV or XLSX file into a list of dicts."""
        suffix = file_path.suffix.lower()

        if suffix == ".csv":
            return self._read_csv(file_path)
        elif suffix in (".xlsx", ".xls"):
            return self._read_xlsx(file_path)
        else:
            raise ValueError(f"Unsupported file format: {suffix}")

    def _read_csv(self, file_path: Path) -> list[dict[str, Any]]:
        """Read a CSV file."""
        rows = []
        with open(file_path, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                # Clean up keys and values, handling None keys/values
                cleaned = {}
                for k, v in row.items():
                    if k is not None:
                        key = k.strip()
                        value = v.strip() if v else ""
                        cleaned[key] = value
                rows.append(cleaned)
        return rows

    def _read_xlsx(self, file_path: Path) -> list[dict[str, Any]]:
        """Read an XLSX file."""
        from openpyxl import load_workbook

        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        rows = []
        headers = None

        for row_num, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_num == 1:
                headers = [str(cell).strip() if cell else "" for cell in row]
                continue

            # Skip empty rows
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            row_dict = {}
            for i, cell in enumerate(row):
                if i < len(headers) and headers[i]:
                    value = str(cell).strip() if cell is not None else ""
                    row_dict[headers[i]] = value
            rows.append(row_dict)

        wb.close()
        return rows

    def _validate_row(
        self,
        row: dict[str, Any],
        row_num: int,
        file_path: Path,
        seen_line_ids: dict[str, int],
    ) -> ValidationResult:
        """Validate a single row."""
        errors: list[ValidationError] = []
        file_str = str(file_path)

        # Get values
        line_id = row.get("line_id", "").strip()
        quarter = row.get("quarter", "").strip().upper()
        year_str = row.get("year", "").strip()
        cn_code = row.get("cn_code", "").strip()
        product_description = row.get("product_description", "").strip()
        country_of_origin = row.get("country_of_origin", "").strip().upper()
        quantity_str = row.get("quantity", "").strip()
        unit = row.get("unit", "").strip().lower()

        # Validate line_id
        if not line_id:
            errors.append(
                InvalidDataTypeError(
                    "line_id",
                    "non-empty string",
                    "(empty)",
                    ErrorLocation(file_str, row_num, "line_id"),
                )
            )
        elif line_id in seen_line_ids:
            errors.append(
                DuplicateLineIdError(
                    line_id,
                    seen_line_ids[line_id],
                    row_num,
                    ErrorLocation(file_str, row_num, "line_id"),
                )
            )
        else:
            seen_line_ids[line_id] = row_num

        if self.fail_fast and errors:
            return ValidationResult(is_valid=False, errors=errors)

        # Validate quarter
        if quarter not in self.VALID_QUARTERS:
            errors.append(
                InvalidEnumValueError(
                    "quarter",
                    quarter,
                    list(self.VALID_QUARTERS),
                    ErrorLocation(file_str, row_num, "quarter"),
                )
            )

        if self.fail_fast and errors:
            return ValidationResult(is_valid=False, errors=errors)

        # Validate year
        try:
            year = int(year_str)
            if year < 2023 or year > 2030:
                errors.append(
                    InvalidDataTypeError(
                        "year",
                        "integer between 2023-2030",
                        year_str,
                        ErrorLocation(file_str, row_num, "year"),
                    )
                )
        except ValueError:
            errors.append(
                InvalidDataTypeError(
                    "year",
                    "integer",
                    year_str,
                    ErrorLocation(file_str, row_num, "year"),
                )
            )
            year = None

        if self.fail_fast and errors:
            return ValidationResult(is_valid=False, errors=errors)

        # Validate CN code
        if len(cn_code) != 8:
            errors.append(
                InvalidCNCodeFormatError(
                    cn_code,
                    ErrorLocation(file_str, row_num, "cn_code"),
                )
            )
        elif not cn_code.isdigit():
            errors.append(
                InvalidDataTypeError(
                    "cn_code",
                    "8-digit number",
                    cn_code,
                    ErrorLocation(file_str, row_num, "cn_code"),
                )
            )
        elif cn_code[:2] not in ("72", "73", "76"):
            category = {
                "25": "Cement",
                "28": "Fertilizers/Hydrogen",
                "31": "Fertilizers",
                "27": "Electricity",
            }.get(cn_code[:2], "unknown")
            errors.append(
                UnsupportedCNCodeError(
                    cn_code,
                    category,
                    ErrorLocation(file_str, row_num, "cn_code"),
                )
            )

        if self.fail_fast and errors:
            return ValidationResult(is_valid=False, errors=errors)

        # Validate product description
        if not product_description:
            errors.append(
                InvalidDataTypeError(
                    "product_description",
                    "non-empty string",
                    "(empty)",
                    ErrorLocation(file_str, row_num, "product_description"),
                )
            )

        if self.fail_fast and errors:
            return ValidationResult(is_valid=False, errors=errors)

        # Validate country code
        if country_of_origin not in VALID_COUNTRY_CODES:
            errors.append(
                InvalidCountryCodeError(
                    country_of_origin,
                    ErrorLocation(file_str, row_num, "country_of_origin"),
                )
            )

        if self.fail_fast and errors:
            return ValidationResult(is_valid=False, errors=errors)

        # Validate quantity
        try:
            quantity = Decimal(quantity_str)
            if quantity <= 0:
                errors.append(
                    NegativeQuantityError(
                        quantity_str,
                        ErrorLocation(file_str, row_num, "quantity"),
                    )
                )
        except (InvalidOperation, ValueError):
            errors.append(
                InvalidDataTypeError(
                    "quantity",
                    "positive number",
                    quantity_str,
                    ErrorLocation(file_str, row_num, "quantity"),
                )
            )
            quantity = None

        if self.fail_fast and errors:
            return ValidationResult(is_valid=False, errors=errors)

        # Validate unit
        if unit not in self.VALID_UNITS:
            errors.append(
                UnknownUnitError(
                    unit,
                    ErrorLocation(file_str, row_num, "unit"),
                )
            )

        if self.fail_fast and errors:
            return ValidationResult(is_valid=False, errors=errors)

        if errors:
            return ValidationResult(is_valid=False, errors=errors)

        # Parse optional fields
        supplier_id = row.get("supplier_id", "").strip() or None
        installation_id = row.get("installation_id", "").strip() or None
        supplier_certificate_ref = row.get("supplier_certificate_ref", "").strip() or None

        supplier_direct = None
        if row.get("supplier_direct_emissions", "").strip():
            try:
                supplier_direct = Decimal(row["supplier_direct_emissions"].strip())
                if supplier_direct < 0:
                    supplier_direct = None
            except (InvalidOperation, ValueError):
                pass

        supplier_indirect = None
        if row.get("supplier_indirect_emissions", "").strip():
            try:
                supplier_indirect = Decimal(row["supplier_indirect_emissions"].strip())
                if supplier_indirect < 0:
                    supplier_indirect = None
            except (InvalidOperation, ValueError):
                pass

        # Create validated line
        line = ImportLineItem(
            line_id=line_id,
            quarter=Quarter(quarter),
            year=year,
            cn_code=cn_code,
            product_description=product_description,
            country_of_origin=country_of_origin,
            quantity=quantity,
            unit=Unit(unit),
            supplier_id=supplier_id,
            installation_id=installation_id,
            supplier_direct_emissions=supplier_direct,
            supplier_indirect_emissions=supplier_indirect,
            supplier_certificate_ref=supplier_certificate_ref,
        )

        return ValidationResult(is_valid=True, validated_lines=[line])

    def _has_field(self, data: dict, field_path: str) -> bool:
        """Check if a nested field exists in a dict."""
        parts = field_path.split(".")
        current = data

        for part in parts:
            if not isinstance(current, dict) or part not in current:
                return False
            current = current[part]

        return current is not None and current != ""
