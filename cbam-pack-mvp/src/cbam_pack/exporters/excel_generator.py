"""
Excel Summary Generator

Generates Excel summary reports for CBAM calculations.
"""

from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from cbam_pack.calculators.emissions_calculator import CalculationResult
from cbam_pack.models import CBAMConfig


class ExcelSummaryGenerator:
    """
    Generates Excel summary reports.

    Creates human-readable Excel summaries with:
    - Summary statistics
    - Line-level details
    - Aggregated views
    - Assumptions log
    """

    # Styles
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True)
    SUBHEADER_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def __init__(self):
        """Initialize the Excel generator."""
        self.wb: Optional[Workbook] = None

    def generate(
        self,
        calc_result: CalculationResult,
        config: CBAMConfig,
        output_path: Path,
    ) -> None:
        """
        Generate Excel summary report.

        Args:
            calc_result: Calculation results
            config: CBAM configuration
            output_path: Path to save the Excel file
        """
        self.wb = Workbook()

        # Remove default sheet
        default_sheet = self.wb.active
        self.wb.remove(default_sheet)

        # Create sheets
        self._create_summary_sheet(calc_result, config)
        self._create_line_details_sheet(calc_result)
        self._create_aggregated_sheet(calc_result)
        self._create_assumptions_sheet(calc_result)

        # Save
        self.wb.save(output_path)

    def _create_summary_sheet(
        self,
        calc_result: CalculationResult,
        config: CBAMConfig,
    ) -> None:
        """Create summary sheet."""
        ws = self.wb.create_sheet("Summary")

        # Title
        ws["A1"] = "CBAM Quarterly Report Summary"
        ws["A1"].font = Font(size=16, bold=True)
        ws.merge_cells("A1:D1")

        # Report info
        ws["A3"] = "Reporting Period:"
        ws["B3"] = f"{config.reporting_period.quarter.value} {config.reporting_period.year}"
        ws["A4"] = "Declarant:"
        ws["B4"] = config.declarant.name
        ws["A5"] = "EORI Number:"
        ws["B5"] = config.declarant.eori_number
        ws["A6"] = "Generated:"
        ws["B6"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Statistics
        stats = calc_result.statistics

        ws["A8"] = "Summary Statistics"
        ws["A8"].font = Font(size=14, bold=True)

        stats_data = [
            ("Total Import Lines", stats.get("total_lines", 0)),
            ("Total Direct Emissions (tCO2e)", f"{stats.get('total_direct_emissions_tco2e', 0):.2f}"),
            ("Total Indirect Emissions (tCO2e)", f"{stats.get('total_indirect_emissions_tco2e', 0):.2f}"),
            ("Total Emissions (tCO2e)", f"{stats.get('total_emissions_tco2e', 0):.2f}"),
            ("", ""),
            ("Lines with Supplier Direct Data", stats.get("lines_with_supplier_direct_data", 0)),
            ("Lines with Supplier Indirect Data", stats.get("lines_with_supplier_indirect_data", 0)),
            ("Lines Using Default Factors", stats.get("lines_using_defaults", 0)),
            ("Default Factor Usage (%)", f"{stats.get('default_usage_percent', 0):.1f}%"),
        ]

        for i, (label, value) in enumerate(stats_data, start=10):
            ws[f"A{i}"] = label
            ws[f"B{i}"] = value
            if label:
                ws[f"A{i}"].font = Font(bold=True)

        # Column widths
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 25

    def _create_line_details_sheet(self, calc_result: CalculationResult) -> None:
        """Create line-level details sheet."""
        ws = self.wb.create_sheet("Line Details")

        # Headers
        headers = [
            "Line ID",
            "Direct Emissions (tCO2e)",
            "Indirect Emissions (tCO2e)",
            "Total Emissions (tCO2e)",
            "Intensity (tCO2e/t)",
            "Direct Method",
            "Indirect Method",
            "Direct Factor Ref",
            "Indirect Factor Ref",
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.BORDER

        # Data rows
        for row, result in enumerate(calc_result.line_results, start=2):
            ws.cell(row=row, column=1, value=result.line_id)
            ws.cell(row=row, column=2, value=float(result.direct_emissions_tco2e))
            ws.cell(row=row, column=3, value=float(result.indirect_emissions_tco2e))
            ws.cell(row=row, column=4, value=float(result.total_emissions_tco2e))
            ws.cell(row=row, column=5, value=float(result.emissions_intensity))
            ws.cell(row=row, column=6, value=result.method_direct.value)
            ws.cell(row=row, column=7, value=result.method_indirect.value)
            ws.cell(row=row, column=8, value=result.factor_direct_ref)
            ws.cell(row=row, column=9, value=result.factor_indirect_ref)

            for col in range(1, 10):
                ws.cell(row=row, column=col).border = self.BORDER

        # Auto-width columns
        for col in range(1, 10):
            ws.column_dimensions[get_column_letter(col)].width = 20

    def _create_aggregated_sheet(self, calc_result: CalculationResult) -> None:
        """Create aggregated results sheet."""
        ws = self.wb.create_sheet("Aggregated")

        if not calc_result.aggregated_results:
            ws["A1"] = "No aggregated results (aggregation policy: PRESERVE_DETAIL)"
            return

        # Headers
        headers = [
            "CN Code",
            "Country of Origin",
            "Quantity (tonnes)",
            "Direct Emissions (tCO2e)",
            "Indirect Emissions (tCO2e)",
            "Total Emissions (tCO2e)",
            "Weighted Intensity",
            "Line Count",
            "Method Used",
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.BORDER

        # Data rows
        for row, agg in enumerate(calc_result.aggregated_results, start=2):
            ws.cell(row=row, column=1, value=agg.cn_code)
            ws.cell(row=row, column=2, value=agg.country_of_origin)
            ws.cell(row=row, column=3, value=float(agg.total_quantity_tonnes))
            ws.cell(row=row, column=4, value=float(agg.total_direct_emissions_tco2e))
            ws.cell(row=row, column=5, value=float(agg.total_indirect_emissions_tco2e))
            ws.cell(row=row, column=6, value=float(agg.total_emissions_tco2e))
            ws.cell(row=row, column=7, value=float(agg.weighted_intensity))
            ws.cell(row=row, column=8, value=agg.line_count)
            ws.cell(row=row, column=9, value=agg.method_used)

            for col in range(1, 10):
                ws.cell(row=row, column=col).border = self.BORDER

        # Auto-width columns
        for col in range(1, 10):
            ws.column_dimensions[get_column_letter(col)].width = 20

    def _create_assumptions_sheet(self, calc_result: CalculationResult) -> None:
        """Create assumptions log sheet."""
        ws = self.wb.create_sheet("Assumptions")

        if not calc_result.assumptions:
            ws["A1"] = "No assumptions recorded"
            return

        # Headers
        headers = [
            "Type",
            "Description",
            "Rationale",
            "Applies To (Lines)",
            "Factor Reference",
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.BORDER

        # Data rows
        for row, assumption in enumerate(calc_result.assumptions, start=2):
            ws.cell(row=row, column=1, value=assumption.type.value)
            ws.cell(row=row, column=2, value=assumption.description)
            ws.cell(row=row, column=3, value=assumption.rationale)
            ws.cell(row=row, column=4, value=", ".join(assumption.applies_to[:20]))
            ws.cell(row=row, column=5, value=assumption.factor_ref or "")

            for col in range(1, 6):
                ws.cell(row=row, column=col).border = self.BORDER
                ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True)

        # Column widths
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 40
        ws.column_dimensions["D"].width = 30
        ws.column_dimensions["E"].width = 25
