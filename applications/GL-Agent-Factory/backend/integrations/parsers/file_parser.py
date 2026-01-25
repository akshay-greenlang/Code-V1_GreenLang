"""
File Parser Module.

This module provides parsers for various file formats commonly used
in emissions and sustainability data exchange.

Supported formats:
- CSV (comma, semicolon, tab delimited)
- Excel (.xlsx, .xls)
- XML (generic, XBRL)
- JSON

Example:
    >>> parser = FileParserFactory.get_parser("emissions.xlsx")
    >>> records = await parser.parse(file_path)
"""

import csv
import io
import json
import logging
from abc import ABC, abstractmethod
from datetime import datetime
from enum import Enum
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

from pydantic import BaseModel, Field

logger = logging.getLogger(__name__)


class FileFormat(str, Enum):
    """Supported file formats."""

    CSV = "csv"
    EXCEL = "excel"
    XML = "xml"
    JSON = "json"
    XBRL = "xbrl"


class ParsedRecord(BaseModel):
    """
    Standardized parsed record format.

    Attributes:
        row_number: Row number in source file
        data: Parsed data dictionary
        errors: List of validation errors
        warnings: List of warnings
    """

    row_number: int = Field(..., description="Row number")
    data: Dict[str, Any] = Field(..., description="Parsed data")
    errors: List[str] = Field(default_factory=list, description="Errors")
    warnings: List[str] = Field(default_factory=list, description="Warnings")
    source_file: Optional[str] = Field(None, description="Source file name")


class ParseResult(BaseModel):
    """
    Result of parsing a file.

    Attributes:
        records: List of parsed records
        total_rows: Total rows in file
        successful_rows: Successfully parsed rows
        failed_rows: Failed rows
        errors: Global errors
        metadata: File metadata
    """

    records: List[ParsedRecord] = Field(..., description="Parsed records")
    total_rows: int = Field(..., description="Total rows")
    successful_rows: int = Field(..., description="Successful rows")
    failed_rows: int = Field(..., description="Failed rows")
    errors: List[str] = Field(default_factory=list, description="Global errors")
    warnings: List[str] = Field(default_factory=list, description="Warnings")
    metadata: Dict[str, Any] = Field(default_factory=dict, description="Metadata")
    parsed_at: datetime = Field(default_factory=datetime.utcnow)


class ColumnMapping(BaseModel):
    """
    Column mapping configuration.

    Attributes:
        source_column: Source column name
        target_field: Target field name
        data_type: Expected data type
        required: Whether field is required
        default_value: Default value if missing
        transform: Transformation function name
    """

    source_column: str = Field(..., description="Source column name")
    target_field: str = Field(..., description="Target field name")
    data_type: str = Field("string", description="Data type")
    required: bool = Field(False, description="Is required")
    default_value: Optional[Any] = Field(None, description="Default value")
    transform: Optional[str] = Field(None, description="Transform function")


class ParserConfig(BaseModel):
    """
    Parser configuration.

    Attributes:
        file_format: File format
        delimiter: CSV delimiter
        encoding: File encoding
        skip_rows: Rows to skip at start
        column_mapping: Column mappings
        date_format: Date format string
        decimal_separator: Decimal separator
    """

    file_format: FileFormat = Field(..., description="File format")
    delimiter: str = Field(",", description="CSV delimiter")
    encoding: str = Field("utf-8", description="File encoding")
    skip_rows: int = Field(0, description="Rows to skip")
    column_mapping: List[ColumnMapping] = Field(
        default_factory=list, description="Column mappings"
    )
    date_format: str = Field("%Y-%m-%d", description="Date format")
    decimal_separator: str = Field(".", description="Decimal separator")
    header_row: int = Field(0, description="Header row index")
    sheet_name: Optional[str] = Field(None, description="Excel sheet name")


class BaseFileParser(ABC):
    """
    Abstract base class for file parsers.

    All file parsers must implement the parse method.
    """

    def __init__(self, config: ParserConfig):
        """
        Initialize parser.

        Args:
            config: Parser configuration
        """
        self.config = config
        self._column_map: Dict[str, ColumnMapping] = {}

        # Build column map for quick lookup
        for mapping in config.column_mapping:
            self._column_map[mapping.source_column.lower()] = mapping

    @abstractmethod
    async def parse(
        self,
        file_path: Union[str, Path],
    ) -> ParseResult:
        """
        Parse a file.

        Args:
            file_path: Path to file

        Returns:
            Parse result with records
        """
        pass

    @abstractmethod
    async def parse_bytes(
        self,
        data: bytes,
        filename: str = "data",
    ) -> ParseResult:
        """
        Parse file from bytes.

        Args:
            data: File content as bytes
            filename: Original filename

        Returns:
            Parse result with records
        """
        pass

    def _apply_mapping(
        self,
        row: Dict[str, Any],
        row_number: int,
    ) -> ParsedRecord:
        """
        Apply column mapping to a row.

        Args:
            row: Raw row data
            row_number: Row number

        Returns:
            Parsed record
        """
        data = {}
        errors = []
        warnings = []

        # If no mappings defined, use raw data
        if not self._column_map:
            return ParsedRecord(
                row_number=row_number,
                data=row,
                errors=[],
                warnings=[],
            )

        for source_col, value in row.items():
            mapping = self._column_map.get(source_col.lower())

            if not mapping:
                # Column not in mapping, skip or include as-is
                data[source_col] = value
                continue

            # Apply transformation
            try:
                transformed = self._transform_value(
                    value,
                    mapping.data_type,
                    mapping.transform,
                )
                data[mapping.target_field] = transformed

            except Exception as e:
                if mapping.required:
                    errors.append(
                        f"Failed to parse {source_col}: {e}"
                    )
                else:
                    warnings.append(
                        f"Using default for {source_col}: {e}"
                    )
                    data[mapping.target_field] = mapping.default_value

        # Check required fields
        for mapping in self.config.column_mapping:
            if mapping.required and mapping.target_field not in data:
                errors.append(f"Missing required field: {mapping.target_field}")

        return ParsedRecord(
            row_number=row_number,
            data=data,
            errors=errors,
            warnings=warnings,
        )

    def _transform_value(
        self,
        value: Any,
        data_type: str,
        transform: Optional[str],
    ) -> Any:
        """
        Transform a value based on data type and transform function.

        Args:
            value: Raw value
            data_type: Target data type
            transform: Optional transform function

        Returns:
            Transformed value
        """
        if value is None or value == "":
            return None

        # Handle string value
        if isinstance(value, str):
            value = value.strip()
            if value == "":
                return None

        # Apply data type conversion
        if data_type == "float":
            if isinstance(value, str):
                value = value.replace(
                    self.config.decimal_separator, "."
                ).replace(",", "")
            return float(value)

        elif data_type == "int":
            return int(float(value))

        elif data_type == "date":
            if isinstance(value, str):
                return datetime.strptime(value, self.config.date_format)
            return value

        elif data_type == "datetime":
            if isinstance(value, str):
                return datetime.fromisoformat(value.replace("Z", "+00:00"))
            return value

        elif data_type == "bool":
            if isinstance(value, str):
                return value.lower() in ("true", "yes", "1", "y")
            return bool(value)

        # Apply custom transforms
        if transform:
            if transform == "uppercase":
                return str(value).upper()
            elif transform == "lowercase":
                return str(value).lower()
            elif transform == "strip":
                return str(value).strip()

        return value


class CSVParser(BaseFileParser):
    """
    CSV file parser.

    Supports:
    - Various delimiters (comma, semicolon, tab)
    - Custom encoding
    - Header row configuration
    - Column mapping
    """

    async def parse(
        self,
        file_path: Union[str, Path],
    ) -> ParseResult:
        """Parse CSV file from path."""
        path = Path(file_path)

        with open(path, "r", encoding=self.config.encoding) as f:
            content = f.read()

        return await self.parse_bytes(
            content.encode(self.config.encoding),
            path.name,
        )

    async def parse_bytes(
        self,
        data: bytes,
        filename: str = "data.csv",
    ) -> ParseResult:
        """Parse CSV from bytes."""
        records = []
        errors = []
        warnings = []

        try:
            content = data.decode(self.config.encoding)
            reader = csv.DictReader(
                io.StringIO(content),
                delimiter=self.config.delimiter,
            )

            # Skip rows if configured
            for _ in range(self.config.skip_rows):
                next(reader, None)

            row_number = self.config.skip_rows + 1
            for row in reader:
                row_number += 1
                try:
                    record = self._apply_mapping(row, row_number)
                    record.source_file = filename
                    records.append(record)
                except Exception as e:
                    errors.append(f"Row {row_number}: {e}")

        except Exception as e:
            errors.append(f"Failed to parse CSV: {e}")
            logger.error(f"CSV parsing error: {e}", exc_info=True)

        successful = len([r for r in records if not r.errors])
        failed = len([r for r in records if r.errors])

        return ParseResult(
            records=records,
            total_rows=len(records),
            successful_rows=successful,
            failed_rows=failed,
            errors=errors,
            warnings=warnings,
            metadata={
                "filename": filename,
                "format": "csv",
                "encoding": self.config.encoding,
                "delimiter": self.config.delimiter,
            },
        )


class ExcelParser(BaseFileParser):
    """
    Excel file parser.

    Supports:
    - .xlsx and .xls formats
    - Multiple sheets
    - Header row configuration
    - Column mapping
    """

    async def parse(
        self,
        file_path: Union[str, Path],
    ) -> ParseResult:
        """Parse Excel file from path."""
        path = Path(file_path)

        with open(path, "rb") as f:
            content = f.read()

        return await self.parse_bytes(content, path.name)

    async def parse_bytes(
        self,
        data: bytes,
        filename: str = "data.xlsx",
    ) -> ParseResult:
        """Parse Excel from bytes."""
        records = []
        errors = []
        warnings = []

        try:
            import openpyxl

            workbook = openpyxl.load_workbook(
                io.BytesIO(data),
                read_only=True,
                data_only=True,
            )

            # Select sheet
            if self.config.sheet_name:
                if self.config.sheet_name not in workbook.sheetnames:
                    errors.append(f"Sheet not found: {self.config.sheet_name}")
                    sheet = workbook.active
                else:
                    sheet = workbook[self.config.sheet_name]
            else:
                sheet = workbook.active

            # Get header row
            header_row_idx = self.config.header_row + 1  # openpyxl is 1-indexed
            headers = []
            for cell in sheet[header_row_idx]:
                headers.append(str(cell.value) if cell.value else f"col_{len(headers)}")

            # Parse data rows
            data_start = header_row_idx + 1 + self.config.skip_rows
            row_number = data_start - 1

            for row in sheet.iter_rows(min_row=data_start, values_only=True):
                row_number += 1

                # Skip empty rows
                if all(cell is None for cell in row):
                    continue

                try:
                    row_dict = dict(zip(headers, row))
                    record = self._apply_mapping(row_dict, row_number)
                    record.source_file = filename
                    records.append(record)
                except Exception as e:
                    errors.append(f"Row {row_number}: {e}")

            workbook.close()

        except ImportError:
            errors.append("openpyxl not installed for Excel parsing")
        except Exception as e:
            errors.append(f"Failed to parse Excel: {e}")
            logger.error(f"Excel parsing error: {e}", exc_info=True)

        successful = len([r for r in records if not r.errors])
        failed = len([r for r in records if r.errors])

        return ParseResult(
            records=records,
            total_rows=len(records),
            successful_rows=successful,
            failed_rows=failed,
            errors=errors,
            warnings=warnings,
            metadata={
                "filename": filename,
                "format": "excel",
                "sheet": self.config.sheet_name or "active",
            },
        )


class XMLParser(BaseFileParser):
    """
    XML file parser.

    Supports:
    - Generic XML
    - XPath-based data extraction
    - Namespace handling
    """

    async def parse(
        self,
        file_path: Union[str, Path],
    ) -> ParseResult:
        """Parse XML file from path."""
        path = Path(file_path)

        with open(path, "rb") as f:
            content = f.read()

        return await self.parse_bytes(content, path.name)

    async def parse_bytes(
        self,
        data: bytes,
        filename: str = "data.xml",
    ) -> ParseResult:
        """Parse XML from bytes."""
        records = []
        errors = []
        warnings = []

        try:
            from lxml import etree

            root = etree.fromstring(data)

            # Find all record elements
            record_xpath = self.config.column_mapping[0].source_column if self.config.column_mapping else "//*"
            elements = root.xpath(record_xpath)

            row_number = 0
            for element in elements:
                row_number += 1

                try:
                    row_dict = {}
                    for mapping in self.config.column_mapping:
                        # Skip the root xpath
                        if mapping == self.config.column_mapping[0]:
                            continue

                        # Extract value using xpath
                        values = element.xpath(mapping.source_column)
                        if values:
                            value = values[0]
                            if hasattr(value, "text"):
                                row_dict[mapping.target_field] = value.text
                            else:
                                row_dict[mapping.target_field] = str(value)

                    record = ParsedRecord(
                        row_number=row_number,
                        data=row_dict,
                        source_file=filename,
                    )
                    records.append(record)

                except Exception as e:
                    errors.append(f"Element {row_number}: {e}")

        except ImportError:
            errors.append("lxml not installed for XML parsing")
        except Exception as e:
            errors.append(f"Failed to parse XML: {e}")
            logger.error(f"XML parsing error: {e}", exc_info=True)

        successful = len([r for r in records if not r.errors])
        failed = len([r for r in records if r.errors])

        return ParseResult(
            records=records,
            total_rows=len(records),
            successful_rows=successful,
            failed_rows=failed,
            errors=errors,
            warnings=warnings,
            metadata={
                "filename": filename,
                "format": "xml",
            },
        )


class JSONParser(BaseFileParser):
    """
    JSON file parser.

    Supports:
    - JSON arrays
    - JSON objects
    - Nested data with JSONPath
    """

    async def parse(
        self,
        file_path: Union[str, Path],
    ) -> ParseResult:
        """Parse JSON file from path."""
        path = Path(file_path)

        with open(path, "r", encoding=self.config.encoding) as f:
            content = f.read()

        return await self.parse_bytes(
            content.encode(self.config.encoding),
            path.name,
        )

    async def parse_bytes(
        self,
        data: bytes,
        filename: str = "data.json",
    ) -> ParseResult:
        """Parse JSON from bytes."""
        records = []
        errors = []
        warnings = []

        try:
            content = data.decode(self.config.encoding)
            json_data = json.loads(content)

            # Handle array or single object
            if isinstance(json_data, list):
                items = json_data
            elif isinstance(json_data, dict):
                # Try common array keys
                for key in ["data", "items", "records", "results"]:
                    if key in json_data and isinstance(json_data[key], list):
                        items = json_data[key]
                        break
                else:
                    items = [json_data]
            else:
                items = []
                errors.append("Invalid JSON structure")

            row_number = 0
            for item in items:
                row_number += 1

                try:
                    record = self._apply_mapping(item, row_number)
                    record.source_file = filename
                    records.append(record)
                except Exception as e:
                    errors.append(f"Item {row_number}: {e}")

        except json.JSONDecodeError as e:
            errors.append(f"Invalid JSON: {e}")
        except Exception as e:
            errors.append(f"Failed to parse JSON: {e}")
            logger.error(f"JSON parsing error: {e}", exc_info=True)

        successful = len([r for r in records if not r.errors])
        failed = len([r for r in records if r.errors])

        return ParseResult(
            records=records,
            total_rows=len(records),
            successful_rows=successful,
            failed_rows=failed,
            errors=errors,
            warnings=warnings,
            metadata={
                "filename": filename,
                "format": "json",
            },
        )


class FileParserFactory:
    """
    Factory for creating file parsers.

    Example:
        >>> parser = FileParserFactory.get_parser("data.csv")
        >>> result = await parser.parse("data.csv")
    """

    EXTENSION_MAP = {
        ".csv": FileFormat.CSV,
        ".tsv": FileFormat.CSV,
        ".xlsx": FileFormat.EXCEL,
        ".xls": FileFormat.EXCEL,
        ".xml": FileFormat.XML,
        ".json": FileFormat.JSON,
    }

    PARSER_MAP = {
        FileFormat.CSV: CSVParser,
        FileFormat.EXCEL: ExcelParser,
        FileFormat.XML: XMLParser,
        FileFormat.JSON: JSONParser,
    }

    @classmethod
    def get_parser(
        cls,
        filename: str,
        config: Optional[ParserConfig] = None,
    ) -> BaseFileParser:
        """
        Get parser for a file.

        Args:
            filename: Filename to determine format
            config: Optional parser configuration

        Returns:
            Appropriate file parser
        """
        # Determine format from extension
        path = Path(filename)
        ext = path.suffix.lower()

        file_format = cls.EXTENSION_MAP.get(ext)
        if not file_format:
            raise ValueError(f"Unsupported file format: {ext}")

        # Create default config if not provided
        if not config:
            config = ParserConfig(file_format=file_format)

            # Set delimiter for TSV
            if ext == ".tsv":
                config.delimiter = "\t"

        # Get parser class
        parser_class = cls.PARSER_MAP.get(file_format)
        if not parser_class:
            raise ValueError(f"No parser for format: {file_format}")

        return parser_class(config)

    @classmethod
    def get_supported_formats(cls) -> List[str]:
        """Get list of supported file extensions."""
        return list(cls.EXTENSION_MAP.keys())
