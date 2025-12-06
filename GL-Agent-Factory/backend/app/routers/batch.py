"""
Batch Processing API

Provides endpoints for bulk calculations:
- CSV/Excel file upload
- Async job processing
- Progress tracking
- Result download (CSV, Excel, JSON)

Example:
    >>> from app.routers.batch import router
    >>> app.include_router(router, prefix="/v1/batch")
"""

import asyncio
import csv
import hashlib
import io
import json
import logging
import tempfile
import uuid
from datetime import datetime, timezone
from enum import Enum
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional

from fastapi import (
    APIRouter,
    BackgroundTasks,
    Depends,
    File,
    Form,
    HTTPException,
    Query,
    Request,
    UploadFile,
    status,
)
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel, Field

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/batch", tags=["Batch Processing"])


# =============================================================================
# Enums and Constants
# =============================================================================


class JobStatus(str, Enum):
    """Batch job status."""

    PENDING = "pending"
    VALIDATING = "validating"
    PROCESSING = "processing"
    COMPLETED = "completed"
    FAILED = "failed"
    CANCELLED = "cancelled"
    PARTIAL = "partial"  # Completed with some errors


class OutputFormat(str, Enum):
    """Output file formats."""

    CSV = "csv"
    EXCEL = "excel"
    JSON = "json"
    PARQUET = "parquet"


class InputFormat(str, Enum):
    """Input file formats."""

    CSV = "csv"
    EXCEL = "excel"
    JSON = "json"


# Constants
MAX_FILE_SIZE_MB = 100
MAX_ROWS_PER_BATCH = 100000
SUPPORTED_EXCEL_EXTENSIONS = {".xlsx", ".xls"}
SUPPORTED_CSV_EXTENSIONS = {".csv", ".tsv"}


# =============================================================================
# Models
# =============================================================================


class BatchJobConfig(BaseModel):
    """Configuration for a batch job."""

    agent_id: str = Field(..., description="Agent to execute for each row")
    version: Optional[str] = Field(None, description="Agent version (latest if not specified)")
    input_mapping: Dict[str, str] = Field(
        default_factory=dict,
        description="Mapping from file columns to agent input fields",
    )
    output_columns: List[str] = Field(
        default_factory=list,
        description="Agent output fields to include in results",
    )
    output_format: OutputFormat = Field(
        OutputFormat.CSV,
        description="Output file format",
    )
    continue_on_error: bool = Field(
        True,
        description="Continue processing if individual rows fail",
    )
    max_concurrent: int = Field(
        10,
        ge=1,
        le=100,
        description="Maximum concurrent executions",
    )
    timeout_per_row_seconds: int = Field(
        30,
        ge=1,
        le=300,
        description="Timeout for each row execution",
    )


class BatchJobCreate(BaseModel):
    """Request to create a batch job."""

    config: BatchJobConfig
    idempotency_key: Optional[str] = Field(
        None,
        description="Idempotency key to prevent duplicate jobs",
    )


class BatchJobRowResult(BaseModel):
    """Result for a single row in batch processing."""

    row_index: int
    status: str
    inputs: Dict[str, Any]
    outputs: Optional[Dict[str, Any]] = None
    error: Optional[str] = None
    execution_id: Optional[str] = None
    duration_ms: Optional[float] = None


class BatchJobProgress(BaseModel):
    """Progress information for a batch job."""

    total_rows: int
    processed_rows: int
    successful_rows: int
    failed_rows: int
    percent_complete: float
    estimated_time_remaining_seconds: Optional[float] = None
    current_rate_per_second: Optional[float] = None


class BatchJobResponse(BaseModel):
    """Response model for batch job."""

    job_id: str
    status: JobStatus
    config: BatchJobConfig
    progress: BatchJobProgress
    input_file_name: str
    input_file_size_bytes: int
    input_file_hash: str
    output_file_url: Optional[str] = None
    error_message: Optional[str] = None
    created_at: datetime
    started_at: Optional[datetime] = None
    completed_at: Optional[datetime] = None
    tenant_id: str
    created_by: Optional[str] = None

    class Config:
        use_enum_values = True


class BatchJobListResponse(BaseModel):
    """Paginated list of batch jobs."""

    data: List[BatchJobResponse]
    meta: Dict[str, Any]


# =============================================================================
# Storage and State Management
# =============================================================================


class BatchJobStore:
    """
    In-memory store for batch jobs.

    In production, this would be backed by Redis/database.
    """

    def __init__(self):
        self._jobs: Dict[str, Dict[str, Any]] = {}
        self._idempotency_keys: Dict[str, str] = {}  # key -> job_id

    async def create_job(
        self,
        job_id: str,
        config: BatchJobConfig,
        file_name: str,
        file_size: int,
        file_hash: str,
        tenant_id: str,
        user_id: Optional[str],
        idempotency_key: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Create a new batch job."""
        # Check idempotency
        if idempotency_key and idempotency_key in self._idempotency_keys:
            existing_job_id = self._idempotency_keys[idempotency_key]
            if existing_job_id in self._jobs:
                return self._jobs[existing_job_id]

        job = {
            "job_id": job_id,
            "status": JobStatus.PENDING.value,
            "config": config.dict(),
            "progress": {
                "total_rows": 0,
                "processed_rows": 0,
                "successful_rows": 0,
                "failed_rows": 0,
                "percent_complete": 0.0,
            },
            "input_file_name": file_name,
            "input_file_size_bytes": file_size,
            "input_file_hash": file_hash,
            "output_file_url": None,
            "error_message": None,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "started_at": None,
            "completed_at": None,
            "tenant_id": tenant_id,
            "created_by": user_id,
            "row_results": [],
        }

        self._jobs[job_id] = job

        if idempotency_key:
            self._idempotency_keys[idempotency_key] = job_id

        return job

    async def get_job(self, job_id: str, tenant_id: str) -> Optional[Dict[str, Any]]:
        """Get a job by ID."""
        job = self._jobs.get(job_id)
        if job and job["tenant_id"] == tenant_id:
            return job
        return None

    async def update_job(self, job_id: str, updates: Dict[str, Any]) -> None:
        """Update a job."""
        if job_id in self._jobs:
            self._jobs[job_id].update(updates)

    async def list_jobs(
        self,
        tenant_id: str,
        status: Optional[JobStatus] = None,
        limit: int = 20,
        offset: int = 0,
    ) -> tuple[List[Dict[str, Any]], int]:
        """List jobs for a tenant."""
        jobs = [
            j for j in self._jobs.values()
            if j["tenant_id"] == tenant_id
            and (status is None or j["status"] == status.value)
        ]

        # Sort by created_at descending
        jobs.sort(key=lambda x: x["created_at"], reverse=True)

        total = len(jobs)
        return jobs[offset:offset + limit], total

    async def delete_job(self, job_id: str) -> None:
        """Delete a job."""
        if job_id in self._jobs:
            del self._jobs[job_id]


# Global store instance
_job_store = BatchJobStore()


# =============================================================================
# File Processing
# =============================================================================


async def parse_csv_file(file_content: bytes) -> List[Dict[str, Any]]:
    """Parse CSV file content to list of dictionaries."""
    content = file_content.decode("utf-8-sig")  # Handle BOM
    reader = csv.DictReader(io.StringIO(content))
    return list(reader)


async def parse_excel_file(file_content: bytes) -> List[Dict[str, Any]]:
    """Parse Excel file content to list of dictionaries."""
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_501_NOT_IMPLEMENTED,
            detail="Excel support requires openpyxl. Install with: pip install openpyxl",
        )

    workbook = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True)
    sheet = workbook.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h) if h else f"column_{i}" for i, h in enumerate(rows[0])]
    data = []

    for row in rows[1:]:
        row_dict = {}
        for i, value in enumerate(row):
            if i < len(headers):
                row_dict[headers[i]] = value
        data.append(row_dict)

    return data


async def parse_json_file(file_content: bytes) -> List[Dict[str, Any]]:
    """Parse JSON file content to list of dictionaries."""
    data = json.loads(file_content.decode("utf-8"))

    if isinstance(data, list):
        return data
    elif isinstance(data, dict) and "data" in data:
        return data["data"]
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="JSON file must contain an array or object with 'data' array",
        )


async def generate_csv_output(results: List[Dict[str, Any]]) -> bytes:
    """Generate CSV output from results."""
    if not results:
        return b""

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=results[0].keys())
    writer.writeheader()
    writer.writerows(results)

    return output.getvalue().encode("utf-8")


async def generate_excel_output(results: List[Dict[str, Any]]) -> bytes:
    """Generate Excel output from results."""
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_501_NOT_IMPLEMENTED,
            detail="Excel support requires openpyxl",
        )

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Results"

    if not results:
        return b""

    # Write headers
    headers = list(results[0].keys())
    for col, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col, value=header)

    # Write data
    for row_idx, row_data in enumerate(results, 2):
        for col_idx, header in enumerate(headers, 1):
            sheet.cell(row=row_idx, column=col_idx, value=row_data.get(header))

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


async def generate_json_output(results: List[Dict[str, Any]]) -> bytes:
    """Generate JSON output from results."""
    return json.dumps({"data": results}, indent=2, default=str).encode("utf-8")


# =============================================================================
# Batch Processing Engine
# =============================================================================


async def process_batch_job(
    job_id: str,
    rows: List[Dict[str, Any]],
    config: BatchJobConfig,
    tenant_id: str,
) -> None:
    """
    Process a batch job asynchronously.

    Args:
        job_id: The job ID
        rows: List of input rows
        config: Job configuration
        tenant_id: Tenant ID
    """
    logger.info(f"Starting batch job {job_id} with {len(rows)} rows")

    # Update job status
    await _job_store.update_job(job_id, {
        "status": JobStatus.PROCESSING.value,
        "started_at": datetime.now(timezone.utc).isoformat(),
        "progress": {
            "total_rows": len(rows),
            "processed_rows": 0,
            "successful_rows": 0,
            "failed_rows": 0,
            "percent_complete": 0.0,
        },
    })

    results = []
    successful = 0
    failed = 0
    start_time = datetime.now(timezone.utc)

    # Create semaphore for concurrency limiting
    semaphore = asyncio.Semaphore(config.max_concurrent)

    async def process_row(row_index: int, row_data: Dict[str, Any]) -> Dict[str, Any]:
        """Process a single row."""
        nonlocal successful, failed

        async with semaphore:
            row_start = datetime.now(timezone.utc)

            try:
                # Map input columns
                inputs = {}
                for agent_field, file_column in config.input_mapping.items():
                    if file_column in row_data:
                        inputs[agent_field] = row_data[file_column]

                # Execute agent (placeholder)
                # TODO: Call actual agent execution service
                await asyncio.sleep(0.1)  # Simulate processing

                # Mock result
                outputs = {
                    "carbon_footprint": 100.0 + row_index,
                    "unit": "tCO2e",
                    "calculation_id": f"calc-{row_index}",
                }

                duration_ms = (datetime.now(timezone.utc) - row_start).total_seconds() * 1000

                result = {
                    "row_index": row_index,
                    "status": "success",
                    "inputs": inputs,
                    "outputs": outputs,
                    "duration_ms": duration_ms,
                }

                successful += 1

            except Exception as e:
                duration_ms = (datetime.now(timezone.utc) - row_start).total_seconds() * 1000

                result = {
                    "row_index": row_index,
                    "status": "error",
                    "inputs": row_data,
                    "error": str(e),
                    "duration_ms": duration_ms,
                }

                failed += 1

                if not config.continue_on_error:
                    raise

            # Update progress
            processed = successful + failed
            elapsed = (datetime.now(timezone.utc) - start_time).total_seconds()
            rate = processed / elapsed if elapsed > 0 else 0
            remaining = (len(rows) - processed) / rate if rate > 0 else None

            await _job_store.update_job(job_id, {
                "progress": {
                    "total_rows": len(rows),
                    "processed_rows": processed,
                    "successful_rows": successful,
                    "failed_rows": failed,
                    "percent_complete": (processed / len(rows)) * 100,
                    "estimated_time_remaining_seconds": remaining,
                    "current_rate_per_second": rate,
                },
            })

            return result

    # Process all rows
    try:
        tasks = [
            process_row(i, row)
            for i, row in enumerate(rows)
        ]
        results = await asyncio.gather(*tasks, return_exceptions=True)

        # Filter out exceptions
        results = [
            r for r in results
            if not isinstance(r, Exception)
        ]

        # Generate output file
        output_data = []
        for result in results:
            if result["status"] == "success":
                row_output = {**result["inputs"], **result.get("outputs", {})}
            else:
                row_output = {**result["inputs"], "error": result.get("error")}
            output_data.append(row_output)

        # Generate output in requested format
        if config.output_format == OutputFormat.CSV:
            output_content = await generate_csv_output(output_data)
            output_ext = ".csv"
        elif config.output_format == OutputFormat.EXCEL:
            output_content = await generate_excel_output(output_data)
            output_ext = ".xlsx"
        else:
            output_content = await generate_json_output(output_data)
            output_ext = ".json"

        # Save output file
        output_path = Path(tempfile.gettempdir()) / f"batch_{job_id}{output_ext}"
        output_path.write_bytes(output_content)

        # Determine final status
        if failed == 0:
            final_status = JobStatus.COMPLETED
        elif successful == 0:
            final_status = JobStatus.FAILED
        else:
            final_status = JobStatus.PARTIAL

        await _job_store.update_job(job_id, {
            "status": final_status.value,
            "completed_at": datetime.now(timezone.utc).isoformat(),
            "output_file_url": f"/v1/batch/jobs/{job_id}/download",
            "row_results": results,
        })

        logger.info(
            f"Batch job {job_id} completed: "
            f"{successful} successful, {failed} failed"
        )

    except Exception as e:
        logger.error(f"Batch job {job_id} failed: {e}")
        await _job_store.update_job(job_id, {
            "status": JobStatus.FAILED.value,
            "completed_at": datetime.now(timezone.utc).isoformat(),
            "error_message": str(e),
        })


# =============================================================================
# API Endpoints
# =============================================================================


@router.post(
    "/jobs",
    response_model=BatchJobResponse,
    status_code=status.HTTP_202_ACCEPTED,
    summary="Create batch job",
    description="Upload a file and create a batch processing job",
)
async def create_batch_job(
    request: Request,
    background_tasks: BackgroundTasks,
    file: UploadFile = File(..., description="CSV, Excel, or JSON file"),
    config: str = Form(..., description="Job configuration as JSON"),
    idempotency_key: Optional[str] = Form(None, description="Idempotency key"),
) -> BatchJobResponse:
    """
    Create a new batch processing job.

    Upload a CSV, Excel, or JSON file and configure the batch job.
    The job will be processed asynchronously in the background.

    Returns:
        Job details with status and progress information
    """
    # Get tenant context
    tenant_id = getattr(request.state, "tenant_id", "dev_tenant")
    user_id = getattr(request.state, "user_id", None)

    # Parse config
    try:
        config_data = json.loads(config)
        job_config = BatchJobConfig(**config_data)
    except json.JSONDecodeError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid JSON in config parameter",
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Invalid config: {e}",
        )

    # Validate file
    if not file.filename:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File name is required",
        )

    file_ext = Path(file.filename).suffix.lower()
    if file_ext not in {".csv", ".xlsx", ".xls", ".json", ".tsv"}:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Unsupported file format: {file_ext}. Use CSV, Excel, or JSON.",
        )

    # Read file content
    file_content = await file.read()
    file_size = len(file_content)

    # Check file size
    if file_size > MAX_FILE_SIZE_MB * 1024 * 1024:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail=f"File size exceeds maximum of {MAX_FILE_SIZE_MB}MB",
        )

    # Calculate file hash
    file_hash = hashlib.sha256(file_content).hexdigest()

    # Parse file
    try:
        if file_ext in {".csv", ".tsv"}:
            rows = await parse_csv_file(file_content)
        elif file_ext in {".xlsx", ".xls"}:
            rows = await parse_excel_file(file_content)
        else:
            rows = await parse_json_file(file_content)
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Failed to parse file: {e}",
        )

    # Validate row count
    if len(rows) > MAX_ROWS_PER_BATCH:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"File has {len(rows)} rows, maximum is {MAX_ROWS_PER_BATCH}",
        )

    if len(rows) == 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File contains no data rows",
        )

    # Create job
    job_id = f"batch-{uuid.uuid4().hex[:12]}"
    job = await _job_store.create_job(
        job_id=job_id,
        config=job_config,
        file_name=file.filename,
        file_size=file_size,
        file_hash=file_hash,
        tenant_id=tenant_id,
        user_id=user_id,
        idempotency_key=idempotency_key,
    )

    # Update total rows
    await _job_store.update_job(job_id, {
        "progress": {
            **job["progress"],
            "total_rows": len(rows),
        },
    })

    # Start background processing
    background_tasks.add_task(
        process_batch_job,
        job_id,
        rows,
        job_config,
        tenant_id,
    )

    # Get updated job
    job = await _job_store.get_job(job_id, tenant_id)

    return BatchJobResponse(
        job_id=job["job_id"],
        status=JobStatus(job["status"]),
        config=BatchJobConfig(**job["config"]),
        progress=BatchJobProgress(**job["progress"]),
        input_file_name=job["input_file_name"],
        input_file_size_bytes=job["input_file_size_bytes"],
        input_file_hash=job["input_file_hash"],
        output_file_url=job.get("output_file_url"),
        error_message=job.get("error_message"),
        created_at=datetime.fromisoformat(job["created_at"]),
        started_at=datetime.fromisoformat(job["started_at"]) if job.get("started_at") else None,
        completed_at=datetime.fromisoformat(job["completed_at"]) if job.get("completed_at") else None,
        tenant_id=job["tenant_id"],
        created_by=job.get("created_by"),
    )


@router.get(
    "/jobs",
    response_model=BatchJobListResponse,
    summary="List batch jobs",
    description="Get paginated list of batch jobs",
)
async def list_batch_jobs(
    request: Request,
    status_filter: Optional[JobStatus] = Query(None, alias="status"),
    limit: int = Query(20, ge=1, le=100),
    offset: int = Query(0, ge=0),
) -> BatchJobListResponse:
    """List batch jobs for the current tenant."""
    tenant_id = getattr(request.state, "tenant_id", "dev_tenant")

    jobs, total = await _job_store.list_jobs(
        tenant_id=tenant_id,
        status=status_filter,
        limit=limit,
        offset=offset,
    )

    return BatchJobListResponse(
        data=[
            BatchJobResponse(
                job_id=job["job_id"],
                status=JobStatus(job["status"]),
                config=BatchJobConfig(**job["config"]),
                progress=BatchJobProgress(**job["progress"]),
                input_file_name=job["input_file_name"],
                input_file_size_bytes=job["input_file_size_bytes"],
                input_file_hash=job["input_file_hash"],
                output_file_url=job.get("output_file_url"),
                error_message=job.get("error_message"),
                created_at=datetime.fromisoformat(job["created_at"]),
                started_at=datetime.fromisoformat(job["started_at"]) if job.get("started_at") else None,
                completed_at=datetime.fromisoformat(job["completed_at"]) if job.get("completed_at") else None,
                tenant_id=job["tenant_id"],
                created_by=job.get("created_by"),
            )
            for job in jobs
        ],
        meta={
            "total": total,
            "limit": limit,
            "offset": offset,
            "has_more": offset + limit < total,
        },
    )


@router.get(
    "/jobs/{job_id}",
    response_model=BatchJobResponse,
    summary="Get batch job",
    description="Get batch job details and status",
)
async def get_batch_job(
    request: Request,
    job_id: str,
) -> BatchJobResponse:
    """Get batch job details."""
    tenant_id = getattr(request.state, "tenant_id", "dev_tenant")

    job = await _job_store.get_job(job_id, tenant_id)
    if not job:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Batch job {job_id} not found",
        )

    return BatchJobResponse(
        job_id=job["job_id"],
        status=JobStatus(job["status"]),
        config=BatchJobConfig(**job["config"]),
        progress=BatchJobProgress(**job["progress"]),
        input_file_name=job["input_file_name"],
        input_file_size_bytes=job["input_file_size_bytes"],
        input_file_hash=job["input_file_hash"],
        output_file_url=job.get("output_file_url"),
        error_message=job.get("error_message"),
        created_at=datetime.fromisoformat(job["created_at"]),
        started_at=datetime.fromisoformat(job["started_at"]) if job.get("started_at") else None,
        completed_at=datetime.fromisoformat(job["completed_at"]) if job.get("completed_at") else None,
        tenant_id=job["tenant_id"],
        created_by=job.get("created_by"),
    )


@router.get(
    "/jobs/{job_id}/progress",
    response_model=BatchJobProgress,
    summary="Get job progress",
    description="Get real-time progress for a batch job",
)
async def get_batch_job_progress(
    request: Request,
    job_id: str,
) -> BatchJobProgress:
    """Get batch job progress."""
    tenant_id = getattr(request.state, "tenant_id", "dev_tenant")

    job = await _job_store.get_job(job_id, tenant_id)
    if not job:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Batch job {job_id} not found",
        )

    return BatchJobProgress(**job["progress"])


@router.get(
    "/jobs/{job_id}/results",
    response_model=List[BatchJobRowResult],
    summary="Get job results",
    description="Get detailed results for each row",
)
async def get_batch_job_results(
    request: Request,
    job_id: str,
    status_filter: Optional[str] = Query(None, alias="status"),
    limit: int = Query(100, ge=1, le=1000),
    offset: int = Query(0, ge=0),
) -> List[BatchJobRowResult]:
    """Get detailed results for a batch job."""
    tenant_id = getattr(request.state, "tenant_id", "dev_tenant")

    job = await _job_store.get_job(job_id, tenant_id)
    if not job:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Batch job {job_id} not found",
        )

    results = job.get("row_results", [])

    # Filter by status
    if status_filter:
        results = [r for r in results if r.get("status") == status_filter]

    # Paginate
    results = results[offset:offset + limit]

    return [BatchJobRowResult(**r) for r in results]


@router.get(
    "/jobs/{job_id}/download",
    summary="Download results",
    description="Download batch job results as a file",
)
async def download_batch_job_results(
    request: Request,
    job_id: str,
    format: Optional[OutputFormat] = Query(None, description="Override output format"),
):
    """Download batch job results as a file."""
    tenant_id = getattr(request.state, "tenant_id", "dev_tenant")

    job = await _job_store.get_job(job_id, tenant_id)
    if not job:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Batch job {job_id} not found",
        )

    if job["status"] not in [JobStatus.COMPLETED.value, JobStatus.PARTIAL.value]:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Job is not completed",
        )

    # Determine format
    output_format = format or OutputFormat(job["config"].get("output_format", "csv"))

    # Generate output
    results = job.get("row_results", [])
    output_data = []
    for result in results:
        if result["status"] == "success":
            row_output = {**result["inputs"], **result.get("outputs", {})}
        else:
            row_output = {**result["inputs"], "error": result.get("error")}
        output_data.append(row_output)

    if output_format == OutputFormat.CSV:
        content = await generate_csv_output(output_data)
        media_type = "text/csv"
        filename = f"batch_{job_id}_results.csv"
    elif output_format == OutputFormat.EXCEL:
        content = await generate_excel_output(output_data)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = f"batch_{job_id}_results.xlsx"
    else:
        content = await generate_json_output(output_data)
        media_type = "application/json"
        filename = f"batch_{job_id}_results.json"

    return StreamingResponse(
        io.BytesIO(content),
        media_type=media_type,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Length": str(len(content)),
        },
    )


@router.post(
    "/jobs/{job_id}/cancel",
    response_model=BatchJobResponse,
    summary="Cancel batch job",
    description="Cancel a running batch job",
)
async def cancel_batch_job(
    request: Request,
    job_id: str,
) -> BatchJobResponse:
    """Cancel a running batch job."""
    tenant_id = getattr(request.state, "tenant_id", "dev_tenant")

    job = await _job_store.get_job(job_id, tenant_id)
    if not job:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Batch job {job_id} not found",
        )

    if job["status"] not in [JobStatus.PENDING.value, JobStatus.PROCESSING.value]:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Cannot cancel job with status {job['status']}",
        )

    await _job_store.update_job(job_id, {
        "status": JobStatus.CANCELLED.value,
        "completed_at": datetime.now(timezone.utc).isoformat(),
    })

    job = await _job_store.get_job(job_id, tenant_id)

    return BatchJobResponse(
        job_id=job["job_id"],
        status=JobStatus(job["status"]),
        config=BatchJobConfig(**job["config"]),
        progress=BatchJobProgress(**job["progress"]),
        input_file_name=job["input_file_name"],
        input_file_size_bytes=job["input_file_size_bytes"],
        input_file_hash=job["input_file_hash"],
        output_file_url=job.get("output_file_url"),
        error_message=job.get("error_message"),
        created_at=datetime.fromisoformat(job["created_at"]),
        started_at=datetime.fromisoformat(job["started_at"]) if job.get("started_at") else None,
        completed_at=datetime.fromisoformat(job["completed_at"]) if job.get("completed_at") else None,
        tenant_id=job["tenant_id"],
        created_by=job.get("created_by"),
    )


@router.delete(
    "/jobs/{job_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    summary="Delete batch job",
    description="Delete a completed batch job and its results",
)
async def delete_batch_job(
    request: Request,
    job_id: str,
) -> None:
    """Delete a batch job."""
    tenant_id = getattr(request.state, "tenant_id", "dev_tenant")

    job = await _job_store.get_job(job_id, tenant_id)
    if not job:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Batch job {job_id} not found",
        )

    if job["status"] in [JobStatus.PENDING.value, JobStatus.PROCESSING.value]:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Cannot delete a running job. Cancel it first.",
        )

    await _job_store.delete_job(job_id)
