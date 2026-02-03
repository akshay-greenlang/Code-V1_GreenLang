"""
GreenLang Report Indexer Lambda Function
Indexes uploaded reports to OpenSearch/Elasticsearch for search functionality.

This function is triggered by S3 PUT events on report uploads and performs:
- Metadata extraction from various file formats (PDF, HTML, Excel, CSV, JSON)
- Text extraction for full-text search
- Indexing to OpenSearch/Elasticsearch
- Search catalog updates
"""

import os
import json
import logging
import tempfile
import hashlib
from typing import Dict, Any, Optional, List
from datetime import datetime
from urllib.parse import unquote_plus
from io import BytesIO

import boto3
from botocore.exceptions import ClientError

# Configure logging
logger = logging.getLogger()
log_level = os.environ.get('LOG_LEVEL', 'INFO')
logger.setLevel(getattr(logging, log_level))

# Environment variables
ENVIRONMENT = os.environ.get('ENVIRONMENT', 'dev')
OPENSEARCH_ENDPOINT = os.environ.get('OPENSEARCH_ENDPOINT', '')
OPENSEARCH_INDEX = os.environ.get('OPENSEARCH_INDEX', 'greenlang-reports')
SUPPORTED_FORMATS = os.environ.get('SUPPORTED_FORMATS', 'pdf,html,xlsx,csv,json').split(',')
EXTRACT_TEXT = os.environ.get('EXTRACT_TEXT', 'true').lower() == 'true'
DLQ_URL = os.environ.get('DLQ_URL', '')
MAX_CONTENT_SIZE_KB = int(os.environ.get('MAX_CONTENT_SIZE_KB', '100'))

# AWS clients
s3_client = boto3.client('s3')
sqs_client = boto3.client('sqs')

# OpenSearch client (lazy initialization)
_opensearch_client = None


def get_opensearch_client():
    """Get or create OpenSearch client."""
    global _opensearch_client

    if _opensearch_client is None and OPENSEARCH_ENDPOINT:
        try:
            from opensearchpy import OpenSearch, RequestsHttpConnection
            from requests_aws4auth import AWS4Auth

            credentials = boto3.Session().get_credentials()
            region = os.environ.get('AWS_REGION', 'us-east-1')

            awsauth = AWS4Auth(
                credentials.access_key,
                credentials.secret_key,
                region,
                'es',
                session_token=credentials.token
            )

            _opensearch_client = OpenSearch(
                hosts=[{'host': OPENSEARCH_ENDPOINT.replace('https://', ''), 'port': 443}],
                http_auth=awsauth,
                use_ssl=True,
                verify_certs=True,
                connection_class=RequestsHttpConnection,
                timeout=30
            )
            logger.info("OpenSearch client initialized")

        except ImportError:
            logger.warning("OpenSearch packages not available")

    return _opensearch_client


class ReportExtractor:
    """Extract metadata and content from various report formats."""

    def __init__(self, bucket: str, key: str):
        self.bucket = bucket
        self.key = key
        self.temp_file: Optional[str] = None
        self.content: bytes = b''

    def extract(self) -> Dict[str, Any]:
        """
        Extract metadata and content from the report.

        Returns:
            Dictionary containing report metadata and optional content preview
        """
        try:
            # Get object metadata
            head_response = s3_client.head_object(Bucket=self.bucket, Key=self.key)

            # Build base metadata
            metadata = {
                'bucket': self.bucket,
                'key': self.key,
                'file_name': os.path.basename(self.key),
                'file_path': self.key,
                'file_size': head_response.get('ContentLength', 0),
                'content_type': head_response.get('ContentType', 'application/octet-stream'),
                'last_modified': head_response.get('LastModified', datetime.utcnow()).isoformat(),
                'etag': head_response.get('ETag', '').strip('"'),
                's3_metadata': head_response.get('Metadata', {}),
                'indexed_at': datetime.utcnow().isoformat(),
                'environment': ENVIRONMENT
            }

            # Extract file extension
            ext = self._get_extension()
            metadata['file_extension'] = ext

            # Download and process content
            self._download_file()

            # Extract format-specific metadata
            if ext == 'pdf':
                metadata.update(self._extract_pdf())
            elif ext == 'html':
                metadata.update(self._extract_html())
            elif ext in ('xlsx', 'xls'):
                metadata.update(self._extract_excel())
            elif ext == 'csv':
                metadata.update(self._extract_csv())
            elif ext == 'json':
                metadata.update(self._extract_json())
            else:
                metadata['format_metadata'] = {}

            # Calculate document hash
            metadata['content_hash'] = self._calculate_hash()

            # Generate document ID
            metadata['doc_id'] = self._generate_doc_id()

            logger.info(f"Extracted metadata for {self.key}")
            return metadata

        finally:
            self._cleanup()

    def _get_extension(self) -> str:
        """Get file extension from key."""
        key_lower = self.key.lower()

        # Handle compound extensions
        if key_lower.endswith('.tar.gz'):
            return 'tar.gz'

        return key_lower.rsplit('.', 1)[-1] if '.' in key_lower else ''

    def _download_file(self) -> None:
        """Download the file for content extraction."""
        try:
            response = s3_client.get_object(Bucket=self.bucket, Key=self.key)
            self.content = response['Body'].read()
            logger.debug(f"Downloaded {len(self.content)} bytes")
        except ClientError as e:
            logger.error(f"Failed to download {self.key}: {e}")
            raise

    def _extract_pdf(self) -> Dict[str, Any]:
        """Extract metadata and content from PDF files."""
        result = {
            'format_metadata': {},
            'content_preview': ''
        }

        try:
            import fitz  # PyMuPDF

            doc = fitz.open(stream=self.content, filetype="pdf")

            # Extract PDF metadata
            result['format_metadata'] = {
                'page_count': len(doc),
                'title': doc.metadata.get('title', ''),
                'author': doc.metadata.get('author', ''),
                'subject': doc.metadata.get('subject', ''),
                'creator': doc.metadata.get('creator', ''),
                'producer': doc.metadata.get('producer', ''),
                'creation_date': doc.metadata.get('creationDate', ''),
                'modification_date': doc.metadata.get('modDate', '')
            }

            # Extract text content
            if EXTRACT_TEXT:
                text_content = []
                max_chars = MAX_CONTENT_SIZE_KB * 1024

                for page in doc:
                    text = page.get_text()
                    text_content.append(text)

                    if sum(len(t) for t in text_content) > max_chars:
                        break

                result['content_preview'] = '\n'.join(text_content)[:max_chars]

            doc.close()
            logger.debug("PDF extraction complete")

        except ImportError:
            logger.warning("PyMuPDF not available - PDF extraction skipped")
        except Exception as e:
            logger.error(f"PDF extraction error: {e}")

        return result

    def _extract_html(self) -> Dict[str, Any]:
        """Extract metadata and content from HTML files."""
        result = {
            'format_metadata': {},
            'content_preview': ''
        }

        try:
            from bs4 import BeautifulSoup

            soup = BeautifulSoup(self.content, 'html.parser')

            # Extract HTML metadata
            title = soup.find('title')
            meta_tags = soup.find_all('meta')

            meta_dict = {}
            for meta in meta_tags:
                name = meta.get('name', meta.get('property', ''))
                content = meta.get('content', '')
                if name and content:
                    meta_dict[name] = content

            result['format_metadata'] = {
                'title': title.string if title else '',
                'meta_tags': meta_dict,
                'link_count': len(soup.find_all('a')),
                'image_count': len(soup.find_all('img')),
                'table_count': len(soup.find_all('table'))
            }

            # Extract text content
            if EXTRACT_TEXT:
                # Remove script and style elements
                for script in soup(["script", "style"]):
                    script.decompose()

                text = soup.get_text(separator=' ', strip=True)
                result['content_preview'] = text[:MAX_CONTENT_SIZE_KB * 1024]

            logger.debug("HTML extraction complete")

        except ImportError:
            logger.warning("BeautifulSoup not available - HTML extraction skipped")
        except Exception as e:
            logger.error(f"HTML extraction error: {e}")

        return result

    def _extract_excel(self) -> Dict[str, Any]:
        """Extract metadata and content from Excel files."""
        result = {
            'format_metadata': {},
            'content_preview': ''
        }

        try:
            import openpyxl

            wb = openpyxl.load_workbook(BytesIO(self.content), read_only=True, data_only=True)

            # Extract workbook metadata
            sheet_info = []
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                sheet_info.append({
                    'name': sheet_name,
                    'max_row': sheet.max_row,
                    'max_column': sheet.max_column
                })

            result['format_metadata'] = {
                'sheet_count': len(wb.sheetnames),
                'sheets': sheet_info,
                'creator': wb.properties.creator or '',
                'created': str(wb.properties.created) if wb.properties.created else '',
                'modified': str(wb.properties.modified) if wb.properties.modified else ''
            }

            # Extract text content (headers and sample data)
            if EXTRACT_TEXT:
                text_content = []

                for sheet_name in wb.sheetnames[:5]:  # Limit to first 5 sheets
                    sheet = wb[sheet_name]
                    text_content.append(f"Sheet: {sheet_name}")

                    for row_idx, row in enumerate(sheet.iter_rows(max_row=100, values_only=True)):
                        row_text = [str(cell) for cell in row if cell is not None]
                        if row_text:
                            text_content.append(' | '.join(row_text))

                        if len('\n'.join(text_content)) > MAX_CONTENT_SIZE_KB * 1024:
                            break

                result['content_preview'] = '\n'.join(text_content)[:MAX_CONTENT_SIZE_KB * 1024]

            wb.close()
            logger.debug("Excel extraction complete")

        except ImportError:
            logger.warning("openpyxl not available - Excel extraction skipped")
        except Exception as e:
            logger.error(f"Excel extraction error: {e}")

        return result

    def _extract_csv(self) -> Dict[str, Any]:
        """Extract metadata and content from CSV files."""
        result = {
            'format_metadata': {},
            'content_preview': ''
        }

        try:
            import csv
            from io import StringIO

            content_str = self.content.decode('utf-8', errors='replace')

            # Detect dialect
            sample = content_str[:8192]
            try:
                dialect = csv.Sniffer().sniff(sample)
                has_header = csv.Sniffer().has_header(sample)
            except csv.Error:
                dialect = csv.excel
                has_header = True

            # Parse CSV
            reader = csv.reader(StringIO(content_str), dialect)
            rows = list(reader)

            # Extract metadata
            result['format_metadata'] = {
                'row_count': len(rows),
                'column_count': len(rows[0]) if rows else 0,
                'has_header': has_header,
                'delimiter': dialect.delimiter,
                'headers': rows[0] if has_header and rows else []
            }

            # Extract content preview
            if EXTRACT_TEXT:
                text_lines = []
                for row in rows[:100]:  # First 100 rows
                    text_lines.append(' | '.join(str(cell) for cell in row))

                    if len('\n'.join(text_lines)) > MAX_CONTENT_SIZE_KB * 1024:
                        break

                result['content_preview'] = '\n'.join(text_lines)[:MAX_CONTENT_SIZE_KB * 1024]

            logger.debug("CSV extraction complete")

        except Exception as e:
            logger.error(f"CSV extraction error: {e}")

        return result

    def _extract_json(self) -> Dict[str, Any]:
        """Extract metadata and content from JSON files."""
        result = {
            'format_metadata': {},
            'content_preview': ''
        }

        try:
            data = json.loads(self.content.decode('utf-8'))

            # Analyze JSON structure
            def analyze_structure(obj, depth=0, max_depth=3):
                if depth >= max_depth:
                    return {'type': type(obj).__name__}

                if isinstance(obj, dict):
                    return {
                        'type': 'object',
                        'keys': list(obj.keys())[:20],
                        'key_count': len(obj)
                    }
                elif isinstance(obj, list):
                    return {
                        'type': 'array',
                        'length': len(obj),
                        'item_type': type(obj[0]).__name__ if obj else None
                    }
                else:
                    return {'type': type(obj).__name__}

            result['format_metadata'] = {
                'structure': analyze_structure(data),
                'is_array': isinstance(data, list),
                'is_object': isinstance(data, dict),
                'root_keys': list(data.keys())[:20] if isinstance(data, dict) else None,
                'array_length': len(data) if isinstance(data, list) else None
            }

            # Content preview
            if EXTRACT_TEXT:
                result['content_preview'] = json.dumps(data, indent=2, default=str)[:MAX_CONTENT_SIZE_KB * 1024]

            logger.debug("JSON extraction complete")

        except json.JSONDecodeError as e:
            logger.error(f"JSON parse error: {e}")
            result['format_metadata'] = {'parse_error': str(e)}

        return result

    def _calculate_hash(self) -> str:
        """Calculate SHA256 hash of content."""
        return hashlib.sha256(self.content).hexdigest()

    def _generate_doc_id(self) -> str:
        """Generate unique document ID."""
        return hashlib.md5(f"{self.bucket}/{self.key}".encode()).hexdigest()

    def _cleanup(self) -> None:
        """Clean up resources."""
        self.content = b''


def index_to_opensearch(metadata: Dict[str, Any]) -> bool:
    """Index document to OpenSearch."""
    client = get_opensearch_client()

    if not client:
        logger.warning("OpenSearch client not available - skipping indexing")
        return False

    try:
        # Prepare document for indexing
        doc = {
            'title': metadata.get('format_metadata', {}).get('title', metadata['file_name']),
            'file_name': metadata['file_name'],
            'file_path': metadata['file_path'],
            'bucket': metadata['bucket'],
            'file_size': metadata['file_size'],
            'file_extension': metadata['file_extension'],
            'content_type': metadata['content_type'],
            'content_preview': metadata.get('content_preview', ''),
            'content_hash': metadata['content_hash'],
            'last_modified': metadata['last_modified'],
            'indexed_at': metadata['indexed_at'],
            'environment': metadata['environment'],
            'format_metadata': metadata.get('format_metadata', {}),
            's3_metadata': metadata.get('s3_metadata', {}),
            'tags': metadata.get('s3_metadata', {}).get('tags', [])
        }

        # Index document
        response = client.index(
            index=OPENSEARCH_INDEX,
            id=metadata['doc_id'],
            body=doc,
            refresh='wait_for'
        )

        logger.info(f"Indexed document {metadata['doc_id']}: {response.get('result', 'unknown')}")
        return response.get('result') in ('created', 'updated')

    except Exception as e:
        logger.error(f"OpenSearch indexing error: {e}")
        return False


def update_s3_metadata(bucket: str, key: str, metadata: Dict[str, Any]) -> None:
    """Update S3 object tags with indexing metadata."""
    try:
        tags = [
            {'Key': 'indexed', 'Value': 'true'},
            {'Key': 'indexed-at', 'Value': metadata['indexed_at'][:19]},
            {'Key': 'doc-id', 'Value': metadata['doc_id']},
            {'Key': 'index-name', 'Value': OPENSEARCH_INDEX}
        ]

        # Get existing tags and merge
        try:
            existing = s3_client.get_object_tagging(Bucket=bucket, Key=key)
            existing_tags = {tag['Key']: tag['Value'] for tag in existing.get('TagSet', [])}
            for tag in tags:
                existing_tags[tag['Key']] = tag['Value']
            tags = [{'Key': k, 'Value': v[:256]} for k, v in existing_tags.items()][:10]
        except ClientError:
            pass

        s3_client.put_object_tagging(
            Bucket=bucket,
            Key=key,
            Tagging={'TagSet': tags}
        )
        logger.debug(f"Updated S3 tags for {key}")

    except ClientError as e:
        logger.warning(f"Failed to update S3 tags: {e}")


def send_to_dlq(record: Dict, error: str) -> None:
    """Send failed message to dead letter queue."""
    if not DLQ_URL:
        return

    try:
        sqs_client.send_message(
            QueueUrl=DLQ_URL,
            MessageBody=json.dumps({
                'original_record': record,
                'error': error,
                'timestamp': datetime.utcnow().isoformat()
            })
        )
        logger.info("Sent failed message to DLQ")
    except ClientError as e:
        logger.error(f"Failed to send to DLQ: {e}")


def process_s3_event(record: Dict) -> Dict:
    """Process a single S3 event record."""
    s3_info = record.get('s3', {})
    bucket = s3_info.get('bucket', {}).get('name', '')
    key = unquote_plus(s3_info.get('object', {}).get('key', ''))

    if not bucket or not key:
        return {'status': 'error', 'message': 'Missing bucket or key'}

    # Check if this is a report
    if not key.startswith('reports/'):
        return {'status': 'skipped', 'message': 'Not a report path'}

    # Check file extension
    ext = key.lower().rsplit('.', 1)[-1] if '.' in key else ''
    if ext not in SUPPORTED_FORMATS:
        return {'status': 'skipped', 'message': f'Unsupported format: {ext}'}

    # Extract metadata
    extractor = ReportExtractor(bucket, key)
    metadata = extractor.extract()

    # Index to OpenSearch
    indexed = index_to_opensearch(metadata)

    # Update S3 metadata
    if indexed:
        update_s3_metadata(bucket, key, metadata)

    return {
        'status': 'indexed' if indexed else 'extracted',
        'bucket': bucket,
        'key': key,
        'doc_id': metadata['doc_id'],
        'indexed': indexed
    }


def lambda_handler(event: Dict[str, Any], context: Any) -> Dict[str, Any]:
    """
    Lambda handler for report indexing.

    Supports both direct S3 events and SQS-wrapped S3 events.
    """
    logger.info(f"Received event: {json.dumps(event)[:1000]}")

    results = []
    batch_item_failures = []

    # Handle SQS events
    if 'Records' in event and event['Records'][0].get('eventSource') == 'aws:sqs':
        for sqs_record in event['Records']:
            message_id = sqs_record.get('messageId', '')
            try:
                body = json.loads(sqs_record.get('body', '{}'))

                # Handle SNS-wrapped S3 events
                if 'Records' in body:
                    s3_records = body['Records']
                elif 'Message' in body:
                    s3_records = json.loads(body['Message']).get('Records', [])
                else:
                    s3_records = [body]

                for s3_record in s3_records:
                    result = process_s3_event(s3_record)
                    results.append(result)

            except Exception as e:
                logger.error(f"Failed to process message {message_id}: {str(e)}")
                batch_item_failures.append({'itemIdentifier': message_id})
                send_to_dlq(sqs_record, str(e))

    # Handle direct S3 events
    elif 'Records' in event:
        for record in event['Records']:
            try:
                result = process_s3_event(record)
                results.append(result)
            except Exception as e:
                logger.error(f"Failed to process S3 event: {str(e)}")
                results.append({'status': 'error', 'message': str(e)})

    # Build response
    response = {
        'statusCode': 200,
        'body': {
            'processed': len(results),
            'indexed': sum(1 for r in results if r.get('indexed')),
            'skipped': sum(1 for r in results if r.get('status') == 'skipped'),
            'errors': sum(1 for r in results if r.get('status') == 'error'),
            'results': results
        }
    }

    if batch_item_failures:
        response['batchItemFailures'] = batch_item_failures

    logger.info(f"Indexing complete: {response['body']}")
    return response
